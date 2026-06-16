"""生成销售订单测试Excel文件 (全中文内容 + 中文文件名)"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reports')

HDR_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='4472C4')
CELL_FONT = Font(name='微软雅黑', size=10)
WRAP = Alignment(vertical='top', wrap_text=True)
CNTR = Alignment(horizontal='center', vertical='top', wrap_text=True)
BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
P0_FILL = PatternFill('solid', fgColor='FFC7CE')
P1_FILL = PatternFill('solid', fgColor='FFEB9C')
P2_FILL = PatternFill('solid', fgColor='C6EFCE')
SEC_FILL = PatternFill('solid', fgColor='D9E2F3')
P_FILL = {'P0':P0_FILL,'P1':P1_FILL,'P2':P2_FILL}

def wc(ws, r, c, v, font=CELL_FONT, align=WRAP, border=BORDER, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font; cell.alignment = align; cell.border = border
    if fill: cell.fill = fill

def hdr(ws, headers, row=1):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def aw(ws, n, mx=55):
    for c in range(1, n+1):
        lens = [len(str(ws.cell(r,c).value or '').split('\n')[0])+2 for r in range(1, ws.max_row+1)]
        ws.column_dimensions[get_column_letter(c)].width = max(10, min(mx, max(lens)))

# ═══════════════════════════════════════════════════════
#  测试用例数据
# ═══════════════════════════════════════════════════════
TC = [
    ('ORD-FUNC-001','【销售订单】页面加载-表头完整性','销售订单','P0',
     '1. 已登录系统（admin/***）\n2. 拥有销售管理权限',
     '1. 导航到「销售管理→销售订单」\n2. 等待页面加载完成\n3. 检查表头列','—',
     '表头包含8列：销售单号、客户名称、产品类型、销售量、车牌号、销售时间、关联合同、操作','功能测试'),
    ('ORD-FUNC-002','【销售订单】页面加载-数据默认展示','销售订单','P0','同 ORD-FUNC-001',
     '1. 进入销售订单页面\n2. 检查表格数据行和分页信息','—',
     '表格默认加载数据行≥1；分页显示"共N条"','功能测试'),
    ('ORD-FUNC-003','【销售订单】搜索-按销售单号精确查询','销售订单','P0','已知存在的销售单号',
     '1. 在「销售单号」输入框输入已知单号\n2. 点击「查询」按钮\n3. 检查搜索结果','销售单号：XS202605290001',
     '搜索结果唯一；结果中销售单号列值=输入值','功能测试'),
    ('ORD-FUNC-004','【销售订单】搜索-按客户名称模糊查询','销售订单','P0','已知存在的客户名称',
     '1. 在「客户名称」输入框输入关键字\n2. 点击「查询」按钮\n3. 检查所有结果行','客户名称：贵州能源',
     '所有结果显示的客户名称包含"贵州能源"；行数≥1','功能测试'),
    ('ORD-FUNC-005','【销售订单】搜索-按产品类型筛选','销售订单','P1','有不同产品类型的订单',
     '1. 点击「产品类型」下拉框\n2. 选择"LNG"\n3. 点击「查询」按钮\n4. 检查所有行的产品类型列','产品类型：LNG',
     '所有结果行的产品类型Tag显示"LNG"；Tag样式为primary（蓝色）','功能测试'),
    ('ORD-FUNC-006','【销售订单】搜索-日期范围筛选','销售订单','P1','存在跨日期范围的订单',
     '1. 输入开始日期：2020-01-01\n2. 输入结束日期：2030-12-31\n3. 点击「查询」按钮',
     '开始日期：2020-01-01\n结束日期：2030-12-31','所有结果的销售时间在日期范围内；结果数≥1','功能测试'),
    ('ORD-FUNC-007','【销售订单】搜索-组合条件查询','销售订单','P1','存在同时满足多个条件的订单',
     '1. 输入客户名称：贵州能源\n2. 输入开始日期：2020-01-01\n3. 输入结束日期：2030-12-31\n4. 点击「查询」按钮',
     '客户名称：贵州能源\n开始日期：2020-01-01\n结束日期：2030-12-31',
     '客户名称含"贵州能源" AND 销售时间在日期范围内；两个条件同时满足','功能测试'),
    ('ORD-FUNC-008','【销售订单】重置-清空搜索条件','销售订单','P1','已输入任意搜索条件',
     '1. 输入搜索条件（任意）\n2. 点击「重置」按钮\n3. 检查搜索条件区域和表格','—',
     '所有输入框清空；下拉框恢复默认；表格恢复全量数据','功能测试'),
    ('ORD-FUNC-009','【销售订单】新增-弹窗打开','销售订单','P1','已登录，有新增销售权限',
     '1. 点击「新增销售」按钮\n2. 检查弹窗','—',
     '弹窗打开；标题显示"新增销售"；包含：关联合同(下拉)、销售量(吨)、销售时间、车牌号、司机姓名等字段','功能测试'),
    ('ORD-FUNC-010','【销售订单】新增-关联合同选择','销售订单','P1','1. 新增弹窗已打开\n2. 系统中存在合同数据',
     '1. 点击「关联合同」下拉框\n2. 输入客户名称关键字\n3. 从下拉列表选择合同','搜索关键字：贵州能源',
     '下拉列表过滤为匹配的合同；选择合同后，客户名称、产品类型、销售单价自动填入（只读）','功能测试'),
    ('ORD-FUNC-011','【销售订单】新增-必填校验','销售订单','P1','新增弹窗已打开，未选择关联合同',
     '1. 不选择关联合同\n2. 直接点击「保存」按钮','—','表单校验失败；提示"请选择关联合同"','功能测试'),
    ('ORD-FUNC-012','【销售订单】新增-完整流程','销售订单','P0','系统中存在可用的合同',
     '1. 点击「新增销售」\n2. 选择关联合同\n3. 输入销售量：10\n4. 输入车牌号：粤B-TEST01\n5. 输入销售时间\n6. 点击「保存」按钮',
     '合同：HT202605280001\n销售量：10\n车牌号：粤B-TEST01',
     'Toast提示"新增成功"；弹窗关闭；表格刷新，新记录可见','功能测试'),
    ('ORD-FUNC-013','【销售订单】详情-查看订单详情','销售订单','P1','表格中存在订单数据',
     '1. 点击某行操作列的「详情」按钮\n2. 检查弹窗内容','—',
     '弹窗打开；显示该订单的详细信息（销售单号、客户、产品、销售量等）','功能测试'),
    ('ORD-FUNC-014','【销售订单】分页-总条数显示','销售订单','P2','页面已加载',
     '1. 检查分页区域','—','显示"共N条"，N与实际数据量一致','功能测试'),
    ('ORD-FUNC-015','【销售订单】分页-下一页翻页','销售订单','P2','总条数 > 每页条数',
     '1. 点击「下一页」按钮\n2. 检查表格数据','—','表格刷新为下一页数据；无重复数据','功能测试'),

    ('ORD-UI-001','【销售订单】UI-产品类型Tag样式(LNG)','销售订单','P2','表格中有LNG类型的订单',
     '1. 定位产品类型列为LNG的行\n2. 检查Tag元素class','—','Tag包含 el-tag--primary 类（蓝色）','UI测试'),
    ('ORD-UI-002','【销售订单】UI-产品类型Tag样式(焦油)','销售订单','P2','表格中有焦油类型的订单',
     '1. 定位产品类型列为焦油的行\n2. 检查Tag元素class','—','Tag包含 el-tag--warning 类（橙色）','UI测试'),
    ('ORD-UI-003','【销售订单】UI-销售量格式','销售订单','P2','表格中有数据',
     '1. 检查销售量列的所有值','—','格式为"数字 t"（如"9 t"、"0.0001 t"）','UI测试'),
    ('ORD-UI-004','【销售订单】UI-弹窗表单布局','销售订单','P2','新增弹窗已打开',
     '1. 检查弹窗中字段排列','—','字段排列：关联合同→销售单号→客户名称→产品类型→销售量(吨)→销售时间→车牌号→司机姓名→取消/保存','UI测试'),

    ('ORD-EDGE-001','【销售订单】边界-搜索不存在的单号','销售订单','P2','页面已加载',
     '1. 输入不存在的销售单号\n2. 点击「查询」按钮','销售单号：NOTEXIST_99999999','表格显示空数据或提示"暂无数据"','边界值测试'),
    ('ORD-EDGE-002','【销售订单】边界-极大日期范围','销售订单','P2','页面已加载',
     '1. 输入开始日期：2000-01-01\n2. 输入结束日期：2099-12-31\n3. 点击「查询」按钮',
     '开始日期：2000-01-01\n结束日期：2099-12-31','正常返回所有数据，无报错','边界值测试'),
    ('ORD-EDGE-003','【销售订单】边界-浮点数精度(0.0001)','销售订单','P2','1. 新增弹窗已打开\n2. 已选择关联合同',
     '1. 输入销售量：0.0001\n2. 填写其他必填字段\n3. 保存\n4. 搜索该订单查看回显','销售量：0.0001',
     '保存成功；回显显示"0.0001 t"，精度不丢失','边界值测试'),
    ('ORD-EDGE-004','【销售订单】边界-浮点数精度(1.5)','销售订单','P2','同 ORD-EDGE-003',
     '1. 输入销售量：1.5\n2. 保存并查看回显','销售量：1.5','回显显示"1.5 t"','边界值测试'),
    ('ORD-EDGE-005','【销售订单】边界-SQL注入防护','销售订单','P2','页面已加载',
     '1. 在销售单号输入框输入 \'; DROP TABLE--\n2. 点击「查询」按钮',
     '销售单号：\'; DROP TABLE--','系统不报错；无SQL注入风险；正常返回空结果','边界值测试'),

    ('ORD-EXCP-001','【销售订单】异常-超卖防护(超大数量)','销售订单','P0','1. 新增弹窗已打开\n2. 已选择关联合同',
     '1. 输入销售量：99999999（远超合同剩余量）\n2. 点击「保存」按钮','销售量：99999999',
     '保存被拦截；提示"超出合同剩余量"或类似错误；数据库不产生新记录','异常测试'),
    ('ORD-EXCP-002','【销售订单】异常-超卖防护(合理数量不误拦)','销售订单','P1','同 ORD-EXCP-001',
     '1. 输入销售量：1（合理值）\n2. 点击「保存」按钮','销售量：1','保存成功；合理数量不被误拦','异常测试'),
    ('ORD-EXCP-003','【销售订单】异常-关联合同为空提交','销售订单','P1','新增弹窗已打开',
     '1. 不选择关联合同\n2. 输入销售量、车牌号等\n3. 点击「保存」按钮','—','提示"请选择关联合同"；保存被拦截','异常测试'),
    ('ORD-EXCP-004','【销售订单】异常-销售量为空提交','销售订单','P1','1. 新增弹窗已打开\n2. 已选择关联合同',
     '1. 不输入销售量\n2. 点击「保存」按钮','—','提示"请输入销售量"；保存被拦截','异常测试'),
    ('ORD-EXCP-005','【销售订单】异常-车牌号为空提交','销售订单','P2','同 ORD-EXCP-004',
     '1. 输入销售量但不输入车牌号\n2. 点击「保存」按钮','—','提示"请输入车牌号"或保存成功（视是否为必填而定）','异常测试'),
    ('ORD-EXCP-006','【销售订单】异常-销售时间为空提交','销售订单','P1','同 ORD-EXCP-004',
     '1. 输入销售量、车牌号但不选择销售时间\n2. 点击「保存」按钮','—','提示"请选择销售时间"；保存被拦截','异常测试'),

    ('ORD-PERM-001','【销售订单】权限-无权限用户访问','销售订单','P2','准备一个无销售管理权限的账号',
     '1. 使用无权限账号登录\n2. 尝试访问销售订单页面','无权限测试账号','页面拒绝访问或菜单中不显示"销售订单"入口；提示权限不足','权限测试'),
    ('ORD-PERM-002','【销售订单】权限-只读用户不可新增','销售订单','P2','准备一个只有查看权限的账号',
     '1. 使用只读账号登录\n2. 检查页面按钮','只读测试账号','"新增销售"按钮不可见或置灰；操作列"详情"按钮可用','权限测试'),

    ('ORD-DATA-001','【销售订单】数据校验-销售单号格式','销售订单','P1','页面已加载',
     '1. 获取第一行销售单号\n2. 验证格式','—','格式：XS + 年月日 + 序号（如 XS202605290001）','数据校验测试'),
    ('ORD-DATA-002','【销售订单】数据校验-销售时间格式','销售订单','P1','页面已加载',
     '1. 获取第一行销售时间\n2. 验证格式','—','格式：YYYY-MM-DD HH:mm:ss（如 2026-05-29 17:43:00）','数据校验测试'),
    ('ORD-DATA-003','【销售订单】数据校验-关联合同号格式','销售订单','P2','页面已加载',
     '1. 获取第一行关联合同\n2. 验证格式','—','格式：HT + 年月日 + 序号（如 HT202605280001）','数据校验测试'),
    ('ORD-DATA-004','【销售订单】数据校验-分页总数与表格一致','销售订单','P1','页面已加载，数据≤10条',
     '1. 获取分页显示的"共N条"\n2. 手动数表格行数','—','分页总条数 ≥ 表格可见行数','数据校验测试'),

    ('ORD-API-001','【销售订单】接口联动-选择合同后客户自动带入','销售订单','P1','1. 新增弹窗已打开\n2. 已选择关联合同',
     '1. 选择关联合同\n2. 检查「客户名称」字段','—','客户名称自动填入（只读灰显）；客户名称与合同关联客户一致；产品类型同步填入','接口联动测试'),
    ('ORD-API-002','【销售订单】接口联动-选择合同后单价自动带入','销售订单','P1','同 ORD-API-001',
     '1. 选择关联合同\n2. 检查「销售单价」字段','—','销售单价自动填入（来自合同单价）','接口联动测试'),
    ('ORD-API-003','【销售订单】接口联动-新增后表格刷新','销售订单','P1','已成功新增订单',
     '1. 新增成功后\n2. 搜索新订单的车牌号或单号\n3. 验证表格中存在新记录','—','新记录出现在表格中；分页总条数+1','接口联动测试'),
]

# ═══════════════════════════════════════════════════════
#  Bug数据
# ═══════════════════════════════════════════════════════

BUGS = [
    ('【已知限制】销售订单-新增销售-关联合同下拉框无法通过Selenium自动化选中','销售订单','提示','P3',
     '1. 测试环境 http://8.136.215.171:8081\n2. 已登录 admin 账号\n3. 使用 Selenium WebDriver 操作',
     '1. Selenium 打开新增销售弹窗\n2. ActionChains 点击「关联合同」下拉框\n3. send_keys 输入客户名称过滤\n4. click 下拉选项\n5. 检查 Vue model-value 是否更新',
     '客户名称：贵州能源集团有限公司\n下拉选项："分页名称_XX - 客户名"',
     '下拉选项被高亮/点击，弹窗关闭。但 Vue 组件内部的 model-value 未被更新。保存时提示"请选择关联合同"',
     '点击选项后，Vue 组件的 model-value 被更新，合同被选中，客户名称/产品类型/单价自动填入',
     '✅ 手工操作正常：鼠标点击下拉→输入→点击选项→客户自动填入','仅影响 Selenium 自动化脚本，不影响用户实际操作',
     'Element Plus filterable Select 使用 Vue Composition API 管理内部状态。Selenium 派发的 click/MouseEvent 事件未能完整触发组件内部事件链路',
     '1. 前端在关联合同 Select 上添加 data-testid\n2. 或暴露 @change 事件 handler\n3. 自动化侧可通过 Vue DevTools API 直接操作组件实例',
     'artifacts/failures/*.png'),
    ('【环境问题】销售订单-页面加载-远程服务器响应慢导致自动化超时','销售订单','提示','P3',
     '1. 测试环境 http://8.136.215.171:8081\n2. 远程服务器负载较高时',
     '1. Selenium 打开销售订单页面\n2. 等待表格渲染\n3. 部分时刻超时 15s','—',
     '页面首次加载超时（需2-3次重试）；Vue 表格渲染延迟；StaleElementReferenceException 偶发',
     '页面在 10s 内完成加载，DOM 稳定','手工操作正常（页面虽慢但能加载完成）',
     '自动化稳定性波动（28/32 ↔ 18/32）；不影响手工操作','远程服务器带宽/CPU限制导致首次访问响应慢',
     '1. 增加自动化超时容忍度（15s→30s）\n2. 或部署本地测试环境','—'),
]

AUTO_BUGS = [
    ('AUTO-01','SalesOrderPage 列索引定义10列，实际页面8列','PageObject错误','✅ 已修复'),
    ('AUTO-02','"新增销售"按钮XPath限定 el-button--primary，实际使用 el-button--success','定位器错误','✅ 已修复'),
    ('AUTO-03','"详情"按钮XPath未限定操作列，匹配到第1列订单号按钮','定位器错误','✅ 已修复'),
    ('AUTO-04','弹窗label文本不匹配："销售量"→"销售量(吨)"、"发货日期"→"销售时间"','PageObject错误','✅ 已修复'),
    ('AUTO-05','弹窗表单结构误判：假设"客户→合同"级联，实际"合同→客户自动带入"','PageObject错误','✅ 已修复'),
    ('AUTO-06','ContractPage.search() 参数名 name→应为 customer_name','API错误','✅ 已修复'),
    ('AUTO-07','get_column_data() 在Vue重渲染时抛 StaleElementReferenceException','竞态条件','✅ 已修复(JS提取)'),
]

IMPROVEMENTS = [
    ('【改进】新增弹窗-关联合同Select组件增加data-testid','优化改进','P2',
     '为新增弹窗"关联合同"filterable Select组件添加 data-testid 属性，提高自动化测试稳定性。'),
    ('【改进】表头列-增加data-testid属性','优化改进','P3',
     '为表格"详情"按钮添加 data-testid（如 data-testid="btn-detail-{orderNo}"），避免XPath误匹配到第1列订单号按钮。'),
]


def gen_test_cases():
    wb = Workbook()
    ws = wb.active
    ws.title = '销售订单测试用例'
    headers = ['用例编号','用例标题','所属模块','优先级','前置条件','测试步骤','测试数据','预期结果','实际结果','测试类别']

    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = '销售订单 — 测试用例表（可直接导入禅道）'
    c.font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    p0 = sum(1 for t in TC if t[3]=='P0')
    p1 = sum(1 for t in TC if t[3]=='P1')
    p2 = sum(1 for t in TC if t[3]=='P2')
    ws.merge_cells('A2:J2')
    ws['A2'].value = f'共 {len(TC)} 条 | P0={p0} | P1={p1} | P2={p2} | 功能测试15 | UI测试4 | 边界值5 | 异常6 | 权限2 | 数据校验4 | 接口联动3'
    ws['A2'].font = Font(name='微软雅黑', size=9, italic=True, color='666666')
    ws.row_dimensions[2].height = 22

    hdr(ws, headers, 3)
    ws.row_dimensions[3].height = 22
    r = 4; prev = ''
    for t in TC:
        cat = t[8]
        if cat != prev:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            wc(ws, r, 1, f'── {cat} ──', Font(name='微软雅黑', size=10, bold=True, color='1F4E79'), fill=SEC_FILL)
            for cc in range(1, 11): ws.cell(r, cc).border = BORDER
            ws.row_dimensions[r].height = 22; r += 1; prev = cat
        vals = [t[0],t[1],t[2],t[3],t[4],t[5],t[6],t[7],'',t[8]]
        for i, v in enumerate(vals, 1):
            al = CNTR if i == 4 else WRAP
            wc(ws, r, i, v, align=al, fill=P_FILL.get(str(v)))
        ws.row_dimensions[r].height = 65
        r += 1

    for i, w in enumerate([16,38,10,8,28,42,28,38,14,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:J{r-1}'

    path = os.path.join(OUT_DIR, '销售订单-测试用例表.xlsx')
    wb.save(path)
    print(f'[OK] {path}  ({len(TC)}条, P0={p0} P1={p1} P2={p2})')


def gen_bug_report():
    wb = Workbook()

    # Sheet1: Bug列表
    ws = wb.active
    ws.title = 'Bug列表'
    headers = ['Bug标题','所属模块','Bug等级','优先级','前置条件','复现步骤','测试数据',
               '实际结果','预期结果','手工验证','影响范围','根本原因','建议方案','附件建议']

    ws.merge_cells('A1:N1')
    ws['A1'].value = '销售订单 — 禅道Bug分析和提交'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:N2')
    ws['A2'].value = f'功能Bug: 0个 | 已知限制: {len(BUGS)}个 | 自动化Bug: {len(AUTO_BUGS)}个(已修复) | 改进任务: {len(IMPROVEMENTS)}个'
    ws['A2'].font = Font(name='微软雅黑', size=9, italic=True, color='666666')
    ws.row_dimensions[2].height = 22

    hdr(ws, headers, 3)
    r = 4
    for b in BUGS:
        for i, v in enumerate(b, 1):
            al = CNTR if i == 4 else WRAP
            wc(ws, r, i, v, align=al)
        ws.row_dimensions[r].height = 120; r += 1
    for i, w in enumerate([30,12,10,8,28,42,28,40,40,14,28,50,40,30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet2: 自动化Bug
    ws2 = wb.create_sheet('自动化Bug(已修复)')
    for i, h in enumerate(['序号','问题描述','类型','修复状态'], 1):
        ws2.cell(1,i,value=h).font=HDR_FONT; ws2.cell(1,i).fill=HDR_FILL; ws2.cell(1,i).border=BORDER
    for i, row in enumerate(AUTO_BUGS, 2):
        for j, v in enumerate(row, 1): wc(ws2, i, j, v)
    ws2.column_dimensions['A'].width=10; ws2.column_dimensions['B'].width=65
    ws2.column_dimensions['C'].width=18; ws2.column_dimensions['D'].width=20

    # Sheet3: 改进任务
    ws3 = wb.create_sheet('改进任务(建议提交禅道)')
    for i, h in enumerate(['任务标题','类型','优先级','描述'], 1):
        ws3.cell(1,i,value=h).font=HDR_FONT; ws3.cell(1,i).fill=HDR_FILL; ws3.cell(1,i).border=BORDER
    for i, row in enumerate(IMPROVEMENTS, 2):
        for j, v in enumerate(row, 1):
            wc(ws3, i, j, v, fill=P_FILL.get(str(v)))
    ws3.column_dimensions['A'].width=55; ws3.column_dimensions['B'].width=12
    ws3.column_dimensions['C'].width=8; ws3.column_dimensions['D'].width=80

    path = os.path.join(OUT_DIR, '销售订单-禅道Bug分析和提交.xlsx')
    wb.save(path)
    print(f'[OK] {path}  (0功能Bug, {len(BUGS)}已知限制, {len(AUTO_BUGS)}自动化Bug已修复, {len(IMPROVEMENTS)}改进任务)')


if __name__ == '__main__':
    os.makedirs(OUT_DIR, exist_ok=True)
    gen_test_cases()
    gen_bug_report()
