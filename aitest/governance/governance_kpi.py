"""
Governance KPI Collector — L4 Measured: 时序指标存储 + 趋势分析。

职责:
  1. 每次审计后自动采集 KPI 数据点
  2. 时序存储到 governance/kpi/ 目录
  3. 趋势计算: 漂移率、合规率、异常率的变化方向
  4. 模块级 + 全局级聚合

用法:
    from aitest.governance.governance_kpi import KPICollector
    kpi = KPICollector()
    kpi.record_audit("state", "equipment", report)
    trends = kpi.get_trends("state", days=30)
"""

import json
from pathlib import Path
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta

WORKSTUDY = Path(__file__).resolve().parent.parent.parent
KPI_DIR = WORKSTUDY / "governance" / "kpi"
KPI_DATA_DIR = KPI_DIR / "timeseries"  # 时序数据子目录
KPI_DIR.mkdir(parents=True, exist_ok=True)
KPI_DATA_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class KPIDataPoint:
    """单个 KPI 数据点。"""
    timestamp: str
    audit_type: str       # "state" | "sop" | "cost"
    module: str
    # State KPI
    drift_count: int = 0
    error_count: int = 0
    warning_count: int = 0
    overall_status: str = "ok"
    # SOP KPI
    compliance_score: float = 1.0
    violation_count: int = 0
    # Cost KPI
    total_cost: float = 0.0
    alert_count: int = 0
    event_count: int = 0


@dataclass
class KPITrend:
    """KPI 趋势分析结果。"""
    metric: str
    current: float
    previous: float
    direction: str       # "up" | "down" | "stable"
    change_pct: float
    status: str          # "improving" | "degrading" | "stable" | "warning"


class KPICollector:
    """L4: Governance KPI 时序采集器。"""

    def __init__(self):
        KPI_DIR.mkdir(parents=True, exist_ok=True)

    def record_audit(self, audit_type: str, module: str, report: dict) -> KPIDataPoint:
        """记录一次审计的 KPI 数据点。"""
        dp = KPIDataPoint(
            timestamp=datetime.now().isoformat(),
            audit_type=audit_type,
            module=module,
        )

        if audit_type == "state":
            dp.drift_count = report.get("drift_count", 0)
            dp.error_count = report.get("error_count", 0)
            dp.warning_count = report.get("warning_count", 0)
            dp.overall_status = report.get("overall_status", "ok")
        elif audit_type == "sop":
            dp.compliance_score = report.get("overall_compliance", 1.0)
            dp.violation_count = report.get("total_violations", 0)
        elif audit_type == "cost":
            dp.total_cost = report.get("total_cost", 0.0)
            dp.alert_count = report.get("alert_count", 0)
            dp.event_count = report.get("total_events", 0)

        # 持久化
        date_str = datetime.now().strftime("%Y-%m-%d")
        daily_file = KPI_DATA_DIR / f"{audit_type}-{module}-{date_str}.jsonl"
        with open(daily_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(asdict(dp), ensure_ascii=False) + "\n")

        return dp

    def get_trends(self, audit_type: str, days: int = 30) -> list[KPITrend]:
        """计算指定审计类型的 KPI 趋势。"""
        cutoff = datetime.now() - timedelta(days=days)
        data_points = self._load_data_points(audit_type, cutoff)

        if len(data_points) < 2:
            return []

        trends = []
        metrics = self._get_metrics_for_type(audit_type)

        for metric in metrics:
            recent = data_points[-5:] if len(data_points) >= 5 else data_points
            older = data_points[:5] if len(data_points) >= 10 else data_points[:len(data_points)//2]

            if not older or not recent:
                continue

            current_val = self._avg_metric(recent, metric)
            previous_val = self._avg_metric(older, metric)

            if previous_val == 0 and current_val == 0:
                direction, change_pct, status = "stable", 0.0, "stable"
            elif previous_val == 0:
                direction, change_pct = "up", 100.0
                status = "warning" if self._metric_is_bad(metric, "up") else "improving"
            else:
                change_pct = round((current_val - previous_val) / previous_val * 100, 1)
                direction = "up" if change_pct > 5 else "down" if change_pct < -5 else "stable"
                status = "degrading" if self._metric_is_bad(metric, direction) else "improving"

            trends.append(KPITrend(
                metric=metric,
                current=round(current_val, 3),
                previous=round(previous_val, 3),
                direction=direction,
                change_pct=change_pct,
                status=status,
            ))

        return trends

    def get_summary(self, days: int = 30) -> dict:
        """获取治理 KPI 总览。"""
        cutoff = datetime.now() - timedelta(days=days)

        state_points = self._load_data_points("state", cutoff)
        sop_points = self._load_data_points("sop", cutoff)
        cost_points = self._load_data_points("cost", cutoff)

        # State Drift Rate
        total_state_audits = len(state_points)
        drifts_found = sum(1 for dp in state_points if dp.drift_count > 0)
        avg_drifts = sum(dp.drift_count for dp in state_points) / max(total_state_audits, 1)

        # SOP Compliance Rate
        total_sop_audits = len(sop_points)
        avg_compliance = sum(dp.compliance_score for dp in sop_points) / max(total_sop_audits, 1)
        violations_found = sum(1 for dp in sop_points if dp.violation_count > 0)

        # Cost Anomaly Rate
        total_cost_audits = len(cost_points)
        anomalies_found = sum(1 for dp in cost_points if dp.alert_count > 0)
        avg_cost = sum(dp.total_cost for dp in cost_points) / max(total_cost_audits, 1)

        # Audit Coverage — 多少模块被审计过
        state_modules = len(set(dp.module for dp in state_points))
        sop_modules = len(set(dp.module for dp in sop_points))

        # Event throughput
        total_governance_events = self._count_recent_events(days)

        return {
            "period": f"{days}d",
            "generated_at": datetime.now().isoformat(),
            "state_drift": {
                "audits": total_state_audits,
                "drift_rate": round(drifts_found / max(total_state_audits, 1) * 100, 1),
                "avg_drifts_per_audit": round(avg_drifts, 1),
                "modules_covered": state_modules,
            },
            "sop_compliance": {
                "audits": total_sop_audits,
                "avg_compliance": round(avg_compliance, 3),
                "violation_rate": round(violations_found / max(total_sop_audits, 1) * 100, 1),
                "modules_covered": sop_modules,
            },
            "cost": {
                "audits": total_cost_audits,
                "anomaly_rate": round(anomalies_found / max(total_cost_audits, 1) * 100, 1),
                "avg_cost_per_period": round(avg_cost, 4),
            },
            "events": {
                "total_governance_events": total_governance_events,
            },
            "trends": {
                "state": [asdict(t) for t in self.get_trends("state", days)],
                "sop": [asdict(t) for t in self.get_trends("sop", days)],
                "cost": [asdict(t) for t in self.get_trends("cost", days)],
            },
        }

    # ── helpers ──

    def _load_data_points(self, audit_type: str, cutoff: datetime) -> list[KPIDataPoint]:
        """加载指定类型的 KPI 数据点。"""
        points = []
        for f in sorted(KPI_DATA_DIR.glob(f"{audit_type}-*.jsonl")):
            try:
                with open(f, "r", encoding="utf-8") as fp:
                    for line in fp:
                        line = line.strip()
                        if not line:
                            continue
                        d = json.loads(line)
                        ts = datetime.fromisoformat(d.get("timestamp", "2000-01-01T00:00:00"))
                        if ts >= cutoff:
                            points.append(KPIDataPoint(**d))
            except Exception:
                pass
        return points

    @staticmethod
    def _get_metrics_for_type(audit_type: str) -> list[str]:
        if audit_type == "state":
            return ["drift_count", "error_count", "warning_count"]
        elif audit_type == "sop":
            return ["compliance_score", "violation_count"]
        elif audit_type == "cost":
            return ["total_cost", "alert_count"]
        return []

    @staticmethod
    def _avg_metric(points: list[KPIDataPoint], metric: str) -> float:
        if not points:
            return 0.0
        return sum(getattr(p, metric, 0) for p in points) / len(points)

    @staticmethod
    def _metric_is_bad(metric: str, direction: str) -> bool:
        """指标上升是否代表退化。"""
        bad_when_up = {"drift_count", "error_count", "warning_count",
                       "violation_count", "total_cost", "alert_count"}
        bad_when_down = {"compliance_score"}
        return (metric in bad_when_up and direction == "up") or \
               (metric in bad_when_down and direction == "down")

    @staticmethod
    def _count_recent_events(days: int) -> int:
        event_dir = WORKSTUDY / "governance" / ".events"
        if not event_dir.exists():
            return 0
        cutoff = datetime.now() - timedelta(days=days)
        count = 0
        for f in event_dir.glob("*.json"):
            try:
                mtime = datetime.fromtimestamp(f.stat().st_mtime)
                if mtime >= cutoff:
                    count += 1
            except Exception:
                pass
        return count


# ══════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════

def run_kpi_summary(days: int = 30, json_output: bool = False) -> dict:
    """运行 KPI 汇总。"""
    collector = KPICollector()
    summary = collector.get_summary(days=days)

    if json_output:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return summary

    print(f"\n{'='*60}")
    print(f"  Governance KPI Summary ({summary['period']})")
    print(f"  Generated: {summary['generated_at'][:19]}")
    print(f"{'='*60}\n")

    sd = summary["state_drift"]
    print(f"  State Drift:     {sd['drift_rate']:.0f}% drift rate "
          f"({sd['avg_drifts_per_audit']:.1f} avg) over {sd['audits']} audits, "
          f"{sd['modules_covered']} modules")

    sc = summary["sop_compliance"]
    print(f"  SOP Compliance:  {sc['avg_compliance']:.0%} avg "
          f"({sc['violation_rate']:.0f}% violation rate) over {sc['audits']} audits")

    co = summary["cost"]
    print(f"  Cost:            ${co['avg_cost_per_period']:.4f} avg, "
          f"{co['anomaly_rate']:.0f}% anomaly rate over {co['audits']} audits")

    ev = summary["events"]
    print(f"  Events:          {ev['total_governance_events']} governance events")

    print(f"\n  Trends:")
    for audit_type in ["state", "sop", "cost"]:
        trends = summary["trends"].get(audit_type, [])
        if trends:
            print(f"    {audit_type}:")
            for t in trends[:5]:
                icon = {"improving": "+", "degrading": "-", "stable": "=", "warning": "!"}.get(t["status"], "?")
                print(f"      {icon} {t['metric']}: {t['previous']:.3f} -> {t['current']:.3f} "
                      f"({t['change_pct']:+.1f}%) [{t['status']}]")

    print()
    return summary


def export_to_excel(output_path: str = None, days: int = 30) -> str:
    """L4: 导出治理 KPI 到 Excel 报表。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    collector = KPICollector()
    summary = collector.get_summary(days=days)

    wb = openpyxl.Workbook()

    # ── Sheet 1: KPI Summary ──
    ws = wb.active
    ws.title = "KPI Summary"

    FONT_NAME = "微软雅黑"
    header_font = Font(name=FONT_NAME, bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    body_font = Font(name=FONT_NAME, size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Governance KPI Report — {summary['period']}"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16)
    ws["A1"].alignment = center_align
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {summary['generated_at'][:19]}"
    ws["A2"].font = Font(name=FONT_NAME, size=10, color="666666")
    ws["A2"].alignment = center_align

    # State Drift
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "State Drift"
    ws[f"A{row}"].font = header_font
    row += 1
    for col, header in enumerate(["Metric", "Value"], 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = header_font_white; c.fill = header_fill; c.border = thin_border
    row += 1
    sd = summary["state_drift"]
    for metric, value in [("Audits", sd["audits"]), ("Drift Rate", f"{sd['drift_rate']}%"),
                           ("Avg Drifts/Audit", sd["avg_drifts_per_audit"]),
                           ("Modules Covered", sd["modules_covered"])]:
        for col, v in enumerate([metric, value], 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = body_font; c.border = thin_border; c.alignment = center_align
        row += 1

    # SOP Compliance
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "SOP Compliance"
    ws[f"A{row}"].font = header_font
    row += 1
    for col, header in enumerate(["Metric", "Value"], 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = header_font_white; c.fill = header_fill; c.border = thin_border
    row += 1
    sc = summary["sop_compliance"]
    for metric, value in [("Audits", sc["audits"]),
                           ("Avg Compliance", f"{sc['avg_compliance']:.1%}"),
                           ("Violation Rate", f"{sc['violation_rate']}%"),
                           ("Modules Covered", sc["modules_covered"])]:
        for col, v in enumerate([metric, value], 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = body_font; c.border = thin_border; c.alignment = center_align
        row += 1

    # Cost
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Cost Governance"
    ws[f"A{row}"].font = header_font
    row += 1
    for col, header in enumerate(["Metric", "Value"], 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = header_font_white; c.fill = header_fill; c.border = thin_border
    row += 1
    co = summary["cost"]
    for metric, value in [("Audits", co["audits"]), ("Anomaly Rate", f"{co['anomaly_rate']}%"),
                           ("Avg Cost/Period", f"${co['avg_cost_per_period']:.4f}")]:
        for col, v in enumerate([metric, value], 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = body_font; c.border = thin_border; c.alignment = center_align
        row += 1

    # Events
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Event Bus"
    ws[f"A{row}"].font = header_font
    row += 1
    ev = summary["events"]
    for col, v in enumerate(["Governance Events (24h)", ev["total_governance_events"]], 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = body_font; c.border = thin_border; c.alignment = center_align

    # ── Sheet 2: Trends ──
    ws2 = wb.create_sheet("Trends")
    row = 1
    for audit_type in ["state", "sop", "cost"]:
        trends = summary["trends"].get(audit_type, [])
        if not trends:
            continue
        ws2.merge_cells(f"A{row}:F{row}")
        ws2[f"A{row}"] = f"{audit_type.upper()} Trends"
        ws2[f"A{row}"].font = header_font
        row += 1
        for col, header in enumerate(["Metric", "Previous", "Current", "Change %", "Direction", "Status"], 1):
            c = ws2.cell(row=row, column=col, value=header)
            c.font = header_font_white; c.fill = header_fill; c.border = thin_border
        row += 1
        for t in trends:
            vals = [t["metric"], round(t["previous"], 3), round(t["current"], 3),
                    f"{t['change_pct']:+.1f}%", t["direction"], t["status"]]
            for col, v in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=col, value=v)
                cell.font = body_font; cell.border = thin_border; cell.alignment = center_align
                if col == 6:
                    colors = {"improving": "C6EFCE", "degrading": "FFC7CE", "stable": "FFEB9C", "warning": "FFC7CE"}
                    cell.fill = PatternFill(start_color=colors.get(v, "FFFFFF"),
                                            end_color=colors.get(v, "FFFFFF"), fill_type="solid")
            row += 1
        row += 2

    # ── Sheet 3: Raw Data Points ──
    ws3 = wb.create_sheet("Raw Data")
    headers = ["Timestamp", "Type", "Module", "Drifts", "Errors", "Warnings",
               "Compliance", "Violations", "Cost", "Alerts"]
    for col, h in enumerate(headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font_white; c.fill = header_fill; c.border = thin_border

    cutoff = datetime.now() - timedelta(days=days)
    row = 2
    for audit_type in ["state", "sop", "cost"]:
        for dp in collector._load_data_points(audit_type, cutoff):
            vals = [dp.timestamp[:19], dp.audit_type, dp.module,
                    dp.drift_count, dp.error_count, dp.warning_count,
                    dp.compliance_score, dp.violation_count,
                    dp.total_cost, dp.alert_count]
            for col, v in enumerate(vals, 1):
                c = ws3.cell(row=row, column=col, value=v)
                c.font = body_font; c.border = thin_border; c.alignment = center_align
            row += 1

    # Column widths
    for ws_name in [ws, ws2, ws3]:
        for col_idx in range(1, 10):
            ws_name.column_dimensions[get_column_letter(col_idx)].width = 18

    if output_path is None:
        kpi_export = KPI_DIR / "governance"
        kpi_export.mkdir(parents=True, exist_ok=True)
        output_path = str(kpi_export / f"governance-kpi-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx")

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import sys
    if "--excel" in sys.argv:
        path = export_to_excel(days=int(sys.argv[1]) if len(sys.argv) > 1 and sys.argv[1].isdigit() else 30)
        print(f"Exported: {path}")
    else:
        days = int(sys.argv[1]) if len(sys.argv) > 1 else 30
        json_out = "--json" in sys.argv
        run_kpi_summary(days=days, json_output=json_out)
