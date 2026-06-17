"""
Test Case Excel Exporter — 测试用例+结果导出。

读取 TEST_CASES.md + allure-results → 生成带分类行+颜色的 Excel 报表。

用法:
    python -m aitest.testing.testcase_exporter sales customer
    aitest testcase export --module=sales --page=customer
"""

import re
import json
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional

WORKSTUDY = Path(__file__).resolve().parent.parent.parent
GOVERNANCE = WORKSTUDY / "governance"
CONTEXT_MODULES = GOVERNANCE / "context" / "projects" / "web-automation" / "modules"
ALLURE_RESULTS = WORKSTUDY / "ZJSN_Test-master526" / "allure-results"
TEST_SCRIPTS = WORKSTUDY / "ZJSN_Test-master526" / "script"
EXPORT_DIR = GOVERNANCE / "kpi" / "testcases"  # 测试用例报表按模块分目录


@dataclass
class TestCaseRow:
    """单条测试用例。"""
    case_id: str          # TC-CUST-001
    test_type: str        # 功能测试/输入校验测试/边界值测试/...
    module: str           # 客户管理模块
    priority: str         # P0/P1/P2
    scenario: str         # 测试场景
    steps: str            # 测试步骤
    test_data: str        # 测试数据
    expected: str         # 预期结果
    actual: str = ""      # 实际结果 (from allure)
    status: str = ""      # pass/fail/broken/skip/untested


# 测试类型颜色映射
TYPE_COLORS = {
    "功能测试":   {"fill": "C6EFCE", "label": "浅绿色"},
    "输入校验测试": {"fill": "BDD7EE", "label": "浅蓝色"},
    "边界值测试":  {"fill": "FFEB9C", "label": "浅黄色"},
    "权限测试":   {"fill": "FFC7CE", "label": "浅红色"},
    "异常测试":   {"fill": "F4B4C2", "label": "浅粉色"},
    "接口测试":   {"fill": "F4CDA5", "label": "浅橙色"},
    "兼容性测试":  {"fill": "D9C2EC", "label": "浅紫色"},
}

# 结果状态颜色
STATUS_COLORS = {
    "pass":     {"fill": "C6EFCE", "text": "通过"},
    "fail":     {"fill": "FFC7CE", "text": "失败"},
    "broken":   {"fill": "FFD9B3", "text": "阻塞"},
    "skip":     {"fill": "FFEB9C", "text": "跳过"},
    "untested": {"fill": "F2F2F2", "text": "未测试"},
}


def _is_header_row(cols: list[str]) -> bool:
    """检测是否为表头行（含"编号""标题""预期"等关键词）。"""
    header_keywords = {"编号", "用例编号", "标题", "测试标题", "步骤", "测试步骤",
                       "预期", "预期结果", "优先级", "前置条件", "测试数据", "自动化"}
    return any(col == kw for col in cols for kw in header_keywords)


def _case_id_valid(cid: str) -> bool:
    """检查是否为合法的用例ID格式。
    支持: TC-EAC-001, TC-TANK-MON-001, TC-PROD-DR-001, GAS-01, TC-CM-001 等。"""
    return bool(re.match(r'^[A-Z]{2,5}(-[A-Z]{2,8}){0,3}-\d{2,4}$', cid))


def parse_test_cases_md(md_path: Path, module_name: str = "客户管理模块") -> list[TestCaseRow]:
    """解析 TEST_CASES.md 提取测试用例。
    支持两种格式：
    A. 标准8栏: | 用例编号 | 标题 | 优先级 | 前置条件 | 测试步骤 | 测试数据 | 预期结果 | 自动化 |
    B. 简要5栏: | 编号 | 标题 | 步骤 | 预期 | 自动化 |  (优先级从段标题提取)
    """
    if not md_path.exists():
        return []

    text = md_path.read_text(encoding="utf-8")
    cases = []

    # 先尝试按优先级分段 (## P0 / ## P1 / ## P2)，无则整文件为一个section
    sections = [(p, c) for s in text.split('\n## ') if re.match(r'^P\d', s)
                for p, c in [(s[0:2], s[2:])]]
    if not sections:
        # 回退：整文件一个section，优先级从表格行内提取
        sections = [("P1", text)]

    for priority, content in sections:

        # 尝试识别表头行来确定列映射
        col_map: dict[str, int] = {}  # field → col_index
        for line in content.split('\n'):
            line = line.strip()
            if not line.startswith('|'):
                continue
            cols = [c.strip() for c in line.split('|')[1:-1]]
            if not cols:
                continue
            if _is_header_row(cols):
                for i, col in enumerate(cols):
                    cl = col.lower()
                    if any(w in cl for w in ["编号", "用例编号", "case"]):
                        col_map["id"] = i
                    elif "标题" in cl:
                        col_map["title"] = i
                    elif "步骤" in cl:
                        col_map["steps"] = i
                    elif "预期" in cl:
                        col_map["expected"] = i
                    elif "优先级" in cl:
                        col_map["priority"] = i
                    elif "前置" in cl:
                        col_map["precondition"] = i
                    elif "数据" in cl:
                        col_map["data"] = i
                continue
            # 数据行
            if not _case_id_valid(cols[0]):
                continue

            n = len(cols)

            if col_map:
                case_id = cols[col_map.get("id", 0)] if col_map.get("id") is not None and col_map["id"] < n else cols[0]
                title = cols[col_map.get("title", 1)] if col_map.get("title") is not None and col_map["title"] < n else ""
                priority_val = cols[col_map.get("priority", 2)] if col_map.get("priority") is not None and col_map["priority"] < n else priority
                precondition = cols[col_map.get("precondition", -1)] if col_map.get("precondition") is not None and col_map["precondition"] < n else ""
                steps = cols[col_map.get("steps", 3)] if col_map.get("steps") is not None and col_map["steps"] < n else ""
                test_data = cols[col_map.get("data", -1)] if col_map.get("data") is not None and col_map["data"] < n else ""
                expected = cols[col_map.get("expected", 4)] if col_map.get("expected") is not None and col_map["expected"] < n else ""
            else:
                # 无表头回退：假设标准8栏
                case_id = cols[0]
                title = cols[1] if n > 1 else ""
                priority_val = cols[2] if n > 2 else priority
                precondition = cols[3] if n > 3 else ""
                steps = cols[4] if n > 4 else ""
                test_data = cols[5] if n > 5 else ""
                expected = cols[6] if n > 6 else ""

            # 分类映射
            test_type = _classify_test_case(case_id, title, steps, expected, priority_val)

            scenario = title
            full_steps = f"{precondition}; {steps}" if precondition else steps

            cases.append(TestCaseRow(
                case_id=case_id,
                test_type=test_type,
                module=module_name,
                priority=priority_val,
                scenario=scenario,
                steps=full_steps,
                test_data=test_data,
                expected=expected,
                status="untested",
            ))

    # ── 回退：非表格格式 (## TC-XXX: Title + bullet list) ──
    if not cases:
        cases = _parse_bullet_format(text, module_name)

    return cases


def _parse_bullet_format(text: str, module_name: str) -> list[TestCaseRow]:
    """解析 ## TC-XXX: Title + bullet list 格式的 TEST_CASES.md。"""
    cases = []
    pattern = re.compile(r'^##\s+(TC-[\w-]+):\s*(.+)$', re.MULTILINE)
    matches = list(pattern.finditer(text))

    for i, m in enumerate(matches):
        case_id = m.group(1).strip()
        title = m.group(2).strip()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[start:end]

        precondition = _extract_bullet_field(body, '前置条件?|前置')
        steps = _extract_bullet_field(body, '步骤|测试步骤')
        expected = _extract_bullet_field(body, '预期|断言')
        priority_val = "P1"
        prio_match = re.search(r'优先级[：:]\s*(P[0-4])', body)
        if prio_match:
            priority_val = prio_match.group(1)

        test_type = _classify_test_case(case_id, title, steps, expected, priority_val)
        full_steps = f"{precondition}; {steps}" if precondition else steps

        cases.append(TestCaseRow(
            case_id=case_id, test_type=test_type, module=module_name,
            priority=priority_val, scenario=title, steps=full_steps,
            test_data="", expected=expected, status="untested",
        ))
    return cases


def _extract_bullet_field(body: str, field_pattern: str) -> str:
    """提取 bullet list 字段内容。"""
    lines = []
    in_field = False
    for line in body.split('\n'):
        stripped = line.strip()
        if re.match(rf'(?:{field_pattern})[：:]', stripped):
            in_field = True
            continue
        if in_field and re.match(r'^[-*+\d.]+\s', stripped):
            lines.append(re.sub(r'^[-*+\d.]+\s*', '', stripped))
        elif in_field and (stripped.startswith('## ') or re.match(r'^(?:预期|断言|步骤|前置|映射)', stripped)):
            in_field = False
    return '; '.join(lines) if lines else "见原文"


def _classify_test_case(case_id: str, title: str, steps: str, expected: str, priority: str) -> str:
    """根据用例内容自动分类到测试类型。"""
    combined = f"{title} {steps} {expected}".lower()

    # 权限相关
    if any(w in combined for w in ["权限", "角色", "销售员", "无权限", "rbac", "admin登录"]):
        return "权限测试"

    # 输入校验
    if any(w in combined for w in ["校验", "格式", "重复", "防重复", "超长", "截断", "必填", "编码重复",
                                     "信用代码", "联系电话", "特殊字符", "xss", "长度"]):
        return "输入校验测试"

    # 边界值
    if any(w in combined for w in ["边界", "分页", "空数据", "无结果", "翻页", "切换", "每页"]):
        return "边界值测试"

    # 异常
    if any(w in combined for w in ["异常", "网络断开", "断网", "离线", "keep-alive", "spa"]):
        return "异常测试"

    # 兼容性
    if any(w in combined for w in ["兼容", "浏览器", "分辨率", "响应式"]):
        return "兼容性测试"

    # 接口
    if any(w in combined for w in ["接口", "api", "请求", "响应", "拦截"]):
        return "接口测试"

    # 默认功能测试
    return "功能测试"


def load_allure_results(page_name: str = None) -> dict[str, dict]:
    """从 allure-results 加载测试结果。返回 {case_name: {status, message}} 映射。"""
    results = {}
    if not ALLURE_RESULTS.exists():
        return results

    for f in ALLURE_RESULTS.glob("*-result.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            name = data.get("name", "")
            status = data.get("status", "unknown")
            status_details = data.get("statusDetails", {})
            message = status_details.get("message", "")[:200] if status_details else ""

            # 筛选指定页面的结果
            if page_name and page_name.lower() not in name.lower():
                continue

            results[name] = {
                "status": status,
                "message": message,
                "fullName": data.get("fullName", name),
            }
        except Exception:
            pass
    return results


def match_results(cases: list[TestCaseRow], allure_results: dict[str, dict]) -> list[TestCaseRow]:
    """将 allure 结果匹配到测试用例。"""
    for case in cases:
        cid = case.case_id.lower()
        matched = None

        for name, result in allure_results.items():
            # 尝试多种匹配策略
            if cid in name.lower():
                matched = result
                break
            # 按标题关键词匹配
            if case.scenario and case.scenario[:4].lower() in name.lower():
                if matched is None:
                    matched = result

        if matched:
            case.status = matched["status"]
            if matched["status"] == "passed":
                case.actual = "✅ 通过"
            elif matched["status"] == "failed":
                case.actual = f"❌ 失败: {matched['message'][:100]}"
            elif matched["status"] == "broken":
                case.actual = f"⚠️ 阻塞: {matched['message'][:100]}"
            elif matched["status"] == "skipped":
                case.actual = "⏭️ 跳过"

    return cases


def export_testcases_to_excel(
    module: str,
    page: str,
    output_path: str = None,
) -> str:
    """导出测试用例+结果到 Excel。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    # 1. 解析 TEST_CASES.md
    md_path = CONTEXT_MODULES / module / "pages" / page / "TEST_CASES.md"
    module_cn = _get_module_cn_name(module, page)
    module_display = f"{module_cn}管理模块"
    cases = parse_test_cases_md(md_path, module_display)

    if not cases:
        raise FileNotFoundError(f"No test cases found in {md_path}")

    # 2. 加载 allure 结果
    allure = load_allure_results(page)
    cases = match_results(cases, allure)

    # 3. 按测试类型分组
    type_order = ["功能测试", "输入校验测试", "边界值测试", "权限测试", "异常测试", "接口测试", "兼容性测试"]
    grouped: dict[str, list[TestCaseRow]] = {t: [] for t in type_order}
    for case in cases:
        if case.test_type in grouped:
            grouped[case.test_type].append(case)
        else:
            grouped.setdefault(case.test_type, []).append(case)

    # 4. 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{module}-{page}"

    # ── 样式 ──
    FONT_NAME = "微软雅黑"
    title_fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")  # 深绿
    title_font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # 深蓝
    header_font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    body_font = Font(name=FONT_NAME, size=10)
    summary_font = Font(name=FONT_NAME, size=10, color="333333")
    cat_font = Font(name=FONT_NAME, bold=True, size=11, color="333333")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    HEADERS = ["用例编号", "测试类型", "所属模块", "优先级", "测试场景", "测试步骤", "测试数据", "预期结果", "实际结果"]
    COL_WIDTHS = [18, 14, 16, 8, 24, 36, 20, 30, 30]

    row = 1

    # ── 标题行 ──
    display_module = cases[0].module if cases else f"{module}/{page}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    title_cell = ws.cell(row=row, column=1, value=f"{display_module}——测试用例表")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    title_cell.border = thin_border
    for c in range(2, len(HEADERS) + 1):
        ws.cell(row=row, column=c).fill = title_fill
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 34
    row += 1

    # ── 统计行 ──
    total = len(cases)
    p0 = sum(1 for c in cases if c.priority == "P0")
    p1 = sum(1 for c in cases if c.priority == "P1")
    p2 = sum(1 for c in cases if c.priority == "P2")
    p3 = sum(1 for c in cases if c.priority == "P3")
    p4 = sum(1 for c in cases if c.priority == "P4")
    passed = sum(1 for c in cases if c.status == "passed")
    failed = sum(1 for c in cases if c.status in ("failed", "broken"))
    untested = sum(1 for c in cases if c.status == "untested")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    summary_cell = ws.cell(row=row, column=1,
                           value=(f"总计: {total}条 | P0: {p0}条 | P1: {p1}条 | P2: {p2}条 | "
                                  f"P3: {p3}条 | P4: {p4}条 | "
                                  f"通过: {passed} | 失败: {failed} | 未测试: {untested} | "
                                  f"通过率: {passed/max(total,1)*100:.1f}%"))
    summary_cell.font = summary_font
    summary_cell.alignment = center_align
    ws.row_dimensions[row].height = 24
    row += 1
    row += 1  # 空行

    # ── 逐分类输出 ──
    case_counter = {}  # type → counter
    for test_type in type_order:
        type_cases = grouped.get(test_type, [])
        if not type_cases:
            continue

        type_fill_color = TYPE_COLORS.get(test_type, {}).get("fill", "F2F2F2")
        type_fill = PatternFill(start_color=type_fill_color, end_color=type_fill_color, fill_type="solid")

        # 分类行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        cat_cell = ws.cell(row=row, column=1,
                           value=f"【{test_type}（{len(type_cases)}条）】")
        cat_cell.font = cat_font
        cat_cell.fill = type_fill
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        cat_cell.border = thin_border
        for c in range(2, len(HEADERS) + 1):
            ws.cell(row=row, column=c).fill = type_fill
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1

        # 表头
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1

        # 用例行
        if test_type not in case_counter:
            # 重新编号
            case_counter = {}
        case_counter.setdefault(test_type, 0)

        for case in type_cases:
            case_counter[test_type] += 1
            seq = case_counter[test_type]

            # 生成新编号 (CUS-FUNC-001 格式)
            type_abbr = _get_type_abbr(test_type)
            new_id = f"CUS-{type_abbr}-{seq:03d}"

            vals = [new_id, test_type, case.module, case.priority,
                    case.scenario, case.steps, case.test_data,
                    case.expected, case.actual or "—"]

            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = center_align

                # 结果列颜色
                if col_idx == 9 and case.status in STATUS_COLORS:
                    sc = STATUS_COLORS[case.status]
                    cell.fill = PatternFill(start_color=sc["fill"], end_color=sc["fill"], fill_type="solid")

            ws.row_dimensions[row].height = 28
            row += 1

        row += 1  # 分类间空行

    # ── 列宽 ──
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 保存: kpi/testcases/{module}/{page}-testcases-{ts}.xlsx ──
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        mod_dir = EXPORT_DIR / module
        mod_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(mod_dir / f"{page}-testcases-{ts}.xlsx")
    else:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _get_module_cn_name(module: str, page: str = "") -> str:
    """模块英文→中文名映射。"""
    mapping = {
        "sales": {"customer": "客户", "contract": "合同", "sales-order": "销售订单", "daily-report": "日报"},
        "personnel": {"employee": "员工", "course": "课程", "exam": "考试", "certificate": "证书",
                      "contractor-unit": "承包商单位", "contractor-personnel": "承包商人员",
                      "entry-approval": "入场审批", "entry-record": "入场记录"},
        "equipment": {"alarm-config": "报警配置", "camera": "摄像头", "key-param": "关键参数", "maintenance": "维保"},
        "system": {"user-list": "用户列表", "user-form": "用户表单", "menu-management": "菜单管理"},
        "production": {"daily-report": "日报", "monthly-report": "月报", "shift-team-config": "班组配置"},
        "tank": {"monitor": "监控", "report": "报表", "alarm-config": "报警配置"},
        "lab": {"gas-analysis-report": "气体分析报告", "gas-compare": "气体对比", "water-report": "水质报告"},
        "warehouse": {},
        "workflow": {},
        "system-role": {"role-list": "角色列表"},
        "dcs": {},
    }
    if module in mapping:
        if page and page in mapping[module]:
            return mapping[module][page]
    # fallback: 拼音翻译
    return page.replace("-", " ").title() if page else module


def _get_type_abbr(test_type: str) -> str:
    """测试类型缩写。"""
    mapping = {
        "功能测试": "FUNC",
        "输入校验测试": "VAL",
        "边界值测试": "BOUND",
        "权限测试": "AUTH",
        "异常测试": "EXC",
        "接口测试": "API",
        "兼容性测试": "COMP",
    }
    return mapping.get(test_type, "OTHER")


# ══════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Usage: python testcase_exporter.py <module> <page> [output.xlsx]")
        print("Example: python testcase_exporter.py sales customer")
        modules_found = []
        if CONTEXT_MODULES.exists():
            for d in sorted(CONTEXT_MODULES.iterdir()):
                if d.is_dir() and not d.name.startswith("."):
                    pages_dir = d / "pages"
                    if pages_dir.exists():
                        pages = [p.name for p in pages_dir.iterdir() if p.is_dir() and not p.name.startswith(".")]
                        if pages:
                            modules_found.append(f"  {d.name}: {', '.join(pages)}")
        if modules_found:
            print("\nAvailable modules/pages:")
            for m in modules_found:
                print(m)
        sys.exit(1)

    module = sys.argv[1]
    page = sys.argv[2]
    output = sys.argv[3] if len(sys.argv) > 3 else None

    path = export_testcases_to_excel(module, page, output)
    print(f"Exported: {path}")
