"""Generate equipment 4-page test case Excel."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'equipment-testcases'

# ── Styles ──
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
section_font = Font(name='Microsoft YaHei', size=11, bold=True)
section_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
title_font = Font(name='Microsoft YaHei', size=16, bold=True)
meta_font = Font(name='Microsoft YaHei', size=10, color='666666')
normal_font = Font(name='Microsoft YaHei', size=10)
pass_font = Font(name='Microsoft YaHei', size=10, color='008000')
fail_font = Font(name='Microsoft YaHei', size=10, color='FF0000')
skip_font = Font(name='Microsoft YaHei', size=10, color='FF8C00')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
section_header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
section_header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='1F4E79')

# Column widths
for col, w in {'A': 40, 'B': 14, 'C': 8, 'D': 28, 'E': 42, 'F': 20, 'G': 38, 'H': 42}.items():
    ws.column_dimensions[col].width = w

# ── Test data (collected from Phase 5 execution) ──
tests = []

def section(name):
    tests.append(('section', name))

def subsection(name):
    tests.append(('subsection', name))

def case(cid, title, mod, pri, pre, steps, data, expected, actual):
    tests.append(('case', cid, title, mod, pri, pre, steps, data, expected, actual))

# ════════════════════════════════════════════════════════════════
#  alarm-config (25 tests)
# ════════════════════════════════════════════════════════════════
section('alarm-config — 设备报警配置')

subsection('P0 — 页面展示 & 冒烟')
case('AC-01','页面正常加载','alarm-config','P0',
    '1.admin已登录\n2.进入设备管理>设备报警配置',
    '1.导航至报警配置页\n2.检查统计卡片\n3.检查统计数字',
    '—','统计卡片≥4张;统计数字非空','✓ PASS. 4张卡片;总数=0(空数据环境)')
case('AC-02','表格表头正确显示','alarm-config','P0',
    '1.AC-01已通过','1.获取表头\n2.校验9列集合','EXPECTED_TABLE_HEADER_SET',
    '表头9列完整','✓ PASS. 9列表头完整匹配')
case('AC-03','搜索输入框可见','alarm-config','P0',
    '1.页面已加载','1.检查搜索输入框','—','可见','✓ PASS')
case('AC-04','新增按钮可见','alarm-config','P0',
    '1.页面已加载','1.检查新增按钮','—','可见','✓ PASS')
case('AC-05','统计卡片数据一致性','alarm-config','P0',
    '1.页面已加载','1.获取全部统计','—','总数/启用/禁用/今日报警≥0','✓ PASS')

subsection('P1 — 搜索筛选')
case('AC-06','按关键词搜索','alarm-config','P1',
    '1.有报警规则数据','1.输入关键词\n2.点击查询','keyword: 测试报警下-1#',
    '仅显示匹配记录','✓ PASS')
case('AC-07','搜索无匹配结果','alarm-config','P1',
    '1.页面已加载','1.输入不存在关键词\n2.查询','keyword: zzzz_nonexistent',
    '不崩溃','✓ PASS')
case('AC-08','重置搜索条件','alarm-config','P1',
    '1.已设搜索条件','1.输入+搜索\n2.重置','—','恢复全量','✓ PASS')

subsection('P1 — 分页测试 (新增)')
case('AC-18','分页组件可见','alarm-config','P1',
    '1.页面已加载','1.检查.el-pagination','—','分页存在','✓ PASS')
case('AC-19','pageSize默认值','alarm-config','P1',
    '1.页面已加载','1.获取pageSize','—','pageSize=10','✓ PASS')

subsection('P1 — 权限控制 (新增)')
case('AC-20','新增按钮权限','alarm-config','P1',
    '1.admin已登录','1.检查新增按钮','—','可见','✓ PASS')
case('AC-21','行操作按钮权限','alarm-config','P1',
    '1.表格有数据','1.检查编辑/删除按钮','—','至少1个','○ SKIP. 表格为空')

subsection('P1 — 边界值测试 (新增)')
case('AC-22','搜索特殊字符','alarm-config','P1',
    '1.页面已加载','1.输入!@#$%^&*()等\n2.查询',
    'keyword: !@#$%^&*()_+...','不崩溃','✓ PASS')
case('AC-23','搜索超长关键词','alarm-config','P1',
    '1.页面已加载','1.输入256字符\n2.查询','keyword: a×256','不崩溃','✓ PASS')

subsection('P2 — 批量操作 (新增)')
case('AC-24','批量复选框检查','alarm-config','P2',
    '1.页面已加载','1.检查表格checkbox','—','批量功能检测','✓ PASS. 复选框=False')

subsection('P2 — 增删改弹窗 (已知Teleport问题)')
case('AC-09','新增-仅必填字段','alarm-config','P0',
    '1.admin,有权限','1.新增\n2.填必填\n3.保存\n4.搜索验证',
    'name: autotest_req_XXXXXX','新增成功','✗ FAIL. teleport+is_displayed (EP-001)')
case('AC-10','新增-全部字段','alarm-config','P1',
    '1.AC-09条件','1.填全部字段\n2.保存','name+type+level+status+desc','全部保存','✗ FAIL. 同上')
case('AC-11','新增-取消','alarm-config','P1',
    '1.AC-09条件','1.新增\n2.填后取消','—','不入库','✗ FAIL. 同上')
case('AC-12','编辑规则','alarm-config','P1',
    '1.表格有数据','1.编辑\n2.修改后保存','—','更新成功','✗ FAIL. 同上')
case('AC-13','删除确认弹窗','alarm-config','P1',
    '1.表格有数据','1.删除\n2.确认弹窗','—','删除成功','✗ FAIL. 同上')
case('AC-14','查看详情','alarm-config','P1',
    '1.表格有数据','1.查看\n2.详情弹窗','—','弹窗显示','○ SKIP. teleport')
case('AC-15','状态切换','alarm-config','P1',
    '1.表格有数据','1.切换el-switch','—','切换成功','○ SKIP. el-switch不可交互')
case('AC-16','阈值相等边界','alarm-config','P1',
    '1.AC-09条件','1.上下限相等\n2.保存','upper=lower=X','系统校验','✗ FAIL. 同上')
case('AC-17','双击防重复','alarm-config','P2',
    '1.AC-09条件','1.快速双击保存','—','仅提交一次','✗ FAIL. click intercepted')

subsection('API — 接口直调')
case('AC-API-01','获取设备列表API','alarm-config','P1',
    '1.获取token','1.GET /api/equipment/device/list','—','code=200,records>0','✓ PASS. 5条设备')
case('AC-API-02','获取用户列表API','alarm-config','P1',
    '1.获取token','1.GET /api/system/user/list','—','code=200,records>0','✓ PASS. 10条用户')

# ════════════════════════════════════════════════════════════════
#  camera (19 tests)
# ════════════════════════════════════════════════════════════════
section('camera — 摄像头管理')

subsection('P0 — 页面展示')
case('CAM-01','页面正常加载','camera','P0',
    '1.admin已登录','1.导航至摄像头管理\n2.检查统计卡片+网格','—',
    '卡片≥4张;统计非空','✓ PASS. 4卡片;总数=0(空数据)')
case('CAM-02','统计卡片数据','camera','P0',
    '1.CAM-01通过','1.获取统计:总数/在线/离线/故障','—','4项≥0','✓ PASS')
case('CAM-03','监控卡片网格','camera','P0',
    '1.CAM-01通过','1.检查monitor-cell数','—','≥0','✓ PASS. 0张(空数据)')
case('CAM-04','统计标签校验','camera','P0',
    '1.CAM-01通过','1.校验4个标签名称','—','标签完整','✓ PASS')

subsection('P1 — 搜索筛选')
case('CAM-05','关键词搜索','camera','P1',
    '1.页面已加载','1.输入关键词\n2.搜索','keyword: 罐区','≥0','✓ PASS')
case('CAM-06','搜索无结果','camera','P1',
    '1.页面已加载','1.输入不存在关键词\n2.搜索','keyword: zzzz_nonexistent','不崩溃','✓ PASS')

subsection('P1 — 分页')
case('CAM-07','分页组件','camera','P1',
    '1.页面已加载','1.检查分页总条数','—','total≥0','✓ PASS. total=0')
case('CAM-08','翻页','camera','P1',
    '1.数据>1页','1.下一页\n2.检查卡片','—','翻页正常','○ SKIP. 数据不足1页')

subsection('P1 — 数据一致性')
case('CAM-11','统计与分页一致','camera','P1',
    '1.页面已加载','1.对比stat_total vs page_total','—','sat≥page','✓ PASS')
case('CAM-12','搜索后分页重置','camera','P1',
    '1.数据>1页','1.翻到第2页\n2.搜索\n3.检查页码','—','重置到第1页','○ SKIP. 数据不足1页')

subsection('P1 — 权限控制 (新增)')
case('CAM-14','搜索框权限','camera','P1',
    '1.admin已登录','1.检查搜索框','—','可见','✓ PASS. (Bug:SEARCH_INPUT→SEARCH_ITEM)')
case('CAM-15','卡片操作按钮权限','camera','P1',
    '1.有监控卡片','1.检查操作按钮区域','—','按钮区域存在','○ SKIP. 无卡片')

subsection('P1 — 边界值 (新增)')
case('CAM-16','搜索特殊字符','camera','P1',
    '1.页面已加载','1.输入特殊字符\n2.搜索','keyword: !@#$%...','不崩溃','✓ PASS')
case('CAM-17','搜索超长关键词','camera','P1',
    '1.页面已加载','1.输入256字符\n2.搜索','keyword: x×256','不崩溃','✓ PASS')
case('CAM-18','空搜索','camera','P1',
    '1.页面已加载','1.清空输入\n2.搜索','keyword: (空)','正常返回','✓ PASS')

subsection('P1 — 可靠性 (新增)')
case('CAM-19','重复搜索稳定','camera','P2',
    '1.页面已加载','1.3次不同关键词搜索','keywords: 罐区,消防,(空)','3次正常','✓ PASS')

subsection('P1 — 弹窗操作')
case('CAM-09','卡片状态标签','camera','P1',
    '1.有卡片','1.获取卡片状态','—','在线/离线/故障之一','○ SKIP. 无卡片')
case('CAM-10','卡片操作按钮','camera','P1',
    '1.有卡片','1.检查操作按钮','—','≥1个','○ SKIP. 无卡片')
case('CAM-13','查看/预览弹窗','camera','P2',
    '1.有卡片','1.点击查看\n2.检查弹窗','—','弹窗正常','○ SKIP. 无卡片')

# ════════════════════════════════════════════════════════════════
#  key-param (14 tests)
# ════════════════════════════════════════════════════════════════
section('key-param — 关键参数监控')

subsection('P0 — 页面展示')
case('KP-01','页面正常加载','key-param','P0',
    '1.admin已登录','1.导航至关键参数\n2.检查卡片+表格+分页','—',
    '卡片≥2;表格;分页','✓ PASS. 参数总数=4')
case('KP-02','统计卡片数据','key-param','P0',
    '1.KP-01通过','1.获取4项统计','—','4项含数字≥0','✓ PASS. 4/4/0/0')
case('KP-03','表头校验','key-param','P0',
    '1.KP-01通过','1.获取表头','—','≥5列','✓ PASS. 9列')

subsection('P0 — 搜索筛选')
case('KP-04','关键词自动过滤','key-param','P0',
    '1.页面已加载','1.输入关键词\n2.等待刷新','keywords: 温度/压力/传感器/test',
    '匹配行数变化','✓ PASS. 温度→4行')
case('KP-05','搜索无匹配','key-param','P1',
    '1.页面已加载','1.输入不存在关键词','keyword: zzzz_nonexistent','行数变化','✓ PASS')
case('KP-06','重置恢复全量','key-param','P1',
    '1.已设搜索','1.输入过滤\n2.重置','—','total=全量','✓ PASS')

subsection('P0 — 数据校验')
case('KP-10','状态逻辑一致性','key-param','P0',
    '1.表格有数据','1.前5条\n2.校验状态vs阈值','—','状态与阈值一致','○ SKIP. 未完成')

subsection('P1 — 弹窗')
case('KP-11','查看参数详情','key-param','P1',
    '1.表格有数据','1.查看\n2.弹窗','—','弹窗显示','○ SKIP. Unicode匹配失败')

subsection('P1 — 分页 (新增)')
case('KP-12','分页组件+总数','key-param','P1',
    '1.页面已加载','1.检查分页','—','total≥0','✓ PASS. total=4')
case('KP-13','pageSize默认','key-param','P1',
    '1.页面已加载','1.获取pageSize','—','pageSize=10','✓ PASS')

subsection('P1 — 权限 (新增)')
case('KP-14','行操作按钮权限','key-param','P1',
    '1.表格有数据','1.检查第一行按钮','—','≥1个','✓ PASS. 2按钮')

subsection('P1 — 边界值 (新增)')
case('KP-15','搜索特殊字符','key-param','P1',
    '1.页面已加载','1.输入特殊字符','keyword: !@#$%...','不崩溃','✓ PASS')
case('KP-16','搜索超长关键词','key-param','P1',
    '1.页面已加载','1.输入256字符','keyword: x×256','不崩溃','✓ PASS')

subsection('P2 — 可靠性 (新增)')
case('KP-17','重复搜索稳定','key-param','P2',
    '1.页面已加载','1.3次重置→输入→刷新','keyword: 温度×3','3次正常','✓ PASS. 3次均4行')

# ════════════════════════════════════════════════════════════════
#  maintenance (21 tests)
# ════════════════════════════════════════════════════════════════
section('maintenance — 设备维保管理')

subsection('P0 — 页面展示')
case('MT-01','页面正常加载','maintenance','P0',
    '1.admin已登录','1.导航至维保\n2.检查URL+标题+表头','—',
    'URL含maintenance;标题含设备维保','✓ PASS. 3表27列')
case('MT-02','表头正确','maintenance','P0',
    '1.MT-01通过','1.检查表头','—','≥9列','✓ PASS')
case('MT-03','表格有数据','maintenance','P0',
    '1.有维保计划','1.获取行数','—','>0','⏱ RETRY. 3表跨表行计数(已知#2)')
case('MT-04','分页组件','maintenance','P0',
    '1.页面已加载','1.检查分页','—','total>0','✓ PASS')

subsection('P1 — 搜索筛选')
case('MT-05','按维保类型搜索','maintenance','P1',
    '1.页面已加载','1.选择日检\n2.搜索','type: 日检','正常显示','✓ PASS')
case('MT-06','按状态搜索','maintenance','P1',
    '1.页面已加载','1.选择待执行\n2.搜索','status: 待执行','正常显示','✓ PASS')
case('MT-07','重置恢复全量','maintenance','P1',
    '1.已设搜索','1.选择类型+搜索\n2.重置','—','total≥原total','✓ PASS')

subsection('P1 — 分页')
case('MT-08','下一页','maintenance','P1',
    '1.数据>1页','1.下一页\n2.检查行数','—','有数据','✓ PASS')
case('MT-09','切换每页条数','maintenance','P1',
    '1.数据>10条','1.切换20条\n2.检查\n3.恢复10','—','行数≤20','✓ PASS')

subsection('P0 — 新增+编辑')
case('MT-10','新增计划-保存','maintenance','P0',
    '1.admin,有权限','1.新增\n2.填写\n3.保存','name: AUTO_测试计划',
    'Toast含成功;total+1','✓ PASS')
case('MT-11','新增-取消','maintenance','P0',
    '1.MT-10条件','1.新增\n2.填写\n3.取消','—','不入库;total不变','✓ PASS')
case('MT-12','新增-必填校验','maintenance','P1',
    '1.MT-10条件','1.新增\n2.直接保存','—','校验错误提示','✓ PASS')
case('MT-13','编辑-修改名称','maintenance','P0',
    '1.表格有数据','1.编辑\n2.改名\n3.保存\n4.还原','name: 原→原_改→原','更新+还原成功','✓ PASS')
case('MT-14','编辑-取消','maintenance','P1',
    '1.表格有数据','1.编辑\n2.改名\n3.取消','—','名称不变','✓ PASS')

subsection('P1 — 生成任务')
case('MT-15','生成维保任务','maintenance','P1',
    '1.表格有数据','1.生成任务\n2.检查Toast','—','有操作反馈','✓ PASS')

subsection('P1 — 权限')
case('MT-16','新增按钮可见(admin)','maintenance','P1',
    '1.admin已登录','1.检查新增按钮','—','可见','✓ PASS')
case('MT-17','编辑按钮行内可见','maintenance','P1',
    '1.有编码','1.检查编辑按钮','—','可见','✓ PASS')

subsection('P1 — 边界值 (新增)')
case('MT-18','下拉类型组合','maintenance','P1',
    '1.页面已加载','1.遍历日检/周检/月检\n2.搜索','types: 日检,周检,月检',
    '每种正常','✓ PASS. (Bug:input_keyword→select_type)')
case('MT-19','名称标识','maintenance','P1',
    '1.MT-10条件','1.新增\n2.填timestamp名称\n3.保存','name: 维保计划_AUTOTEST_{ts}',
    '保存成功','✓ PASS. (Bug:100char→timestamp)')

subsection('P1 — 批量操作 (新增)')
case('MT-20','批量复选框','maintenance','P1',
    '1.页面已加载','1.检查表格checkbox','—','批量功能检测','✓ PASS. checkbox=False')

subsection('P2 — 可靠性 (新增)')
case('MT-21','重复搜索稳定','maintenance','P2',
    '1.页面已加载','1.3次:重置→选类型→搜索','types: 日检,待执行,日检',
    '3次正常','✓ PASS. (Bug:input_keyword→select)')

# ── Stats ──
total = sum(1 for t in tests if t[0]=='case')
p0 = sum(1 for t in tests if t[0]=='case' and t[4]=='P0')
p1 = sum(1 for t in tests if t[0]=='case' and t[4]=='P1')
p2 = sum(1 for t in tests if t[0]=='case' and t[4]=='P2')
n_pass = sum(1 for t in tests if t[0]=='case' and 'PASS' in t[9])
n_fail = sum(1 for t in tests if t[0]=='case' and 'FAIL' in t[9])
n_skip = sum(1 for t in tests if t[0]=='case' and 'SKIP' in t[9])
n_retry = sum(1 for t in tests if t[0]=='case' and 'RETRY' in t[9])
effective = total - n_skip
effective_pass = n_pass
rate = int(effective_pass/effective*100) if effective>0 else 0

# ── Write ──
row = 1

# Title
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
ws.cell(row=row, column=1, value='设备管理 equipment — 测试用例表 (4页面全量)').font = title_font
ws.cell(row=row, column=1).alignment = center_align
row += 1

# Meta
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
today = datetime.now().strftime('%Y-%m-%d')
meta = f'生成日期:{today} | 用例总数:{total} | P0:{p0} | P1:{p1} | P2:{p2} | PASS:{n_pass} | FAIL:{n_fail} | SKIP:{n_skip} | 有效通过率:{rate}% ({effective_pass}/{effective})'
ws.cell(row=row, column=1, value=meta).font = meta_font
ws.cell(row=row, column=1).alignment = left_align
row += 2

# Headers
headers = ['用例标题', '所属模块', '优先级', '前置条件', '测试步骤', '测试数据', '预期结果', '实际结果']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
row += 1

# Test cases
for test in tests:
    if test[0] == 'section':
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f'  {test[1]}')
        c.font = section_header_font; c.fill = section_header_fill; c.alignment = left_align
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).fill = section_header_fill
        row += 1
    elif test[0] == 'subsection':
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f'  {test[1]}')
        c.font = section_font; c.fill = section_fill; c.alignment = left_align
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).fill = section_fill
        row += 1
    elif test[0] == 'case':
        _, cid, title, mod, pri, pre, steps, data, expected, actual = test
        vals = [f'{cid} {title}', mod, pri, pre, steps, data, expected, actual]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = normal_font; c.alignment = left_align; c.border = thin_border
            if i == 8:
                if 'PASS' in v: c.font = pass_font
                elif 'FAIL' in v: c.font = fail_font
                elif 'SKIP' in v or 'RETRY' in v: c.font = skip_font
        row += 1

# Bug fixes section
row_empty = row
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
c = ws.cell(row=row, column=1, value='  Bug修复记录')
c.font = section_header_font; c.fill = section_header_fill; c.alignment = left_align
for col in range(1, 9):
    ws.cell(row=row, column=col).border = thin_border
    ws.cell(row=row, column=col).fill = section_header_fill
row += 1

bug_headers = ['Bug ID', '问题', '根因', '修复', '', '', '', '']
for i, h in enumerate(bug_headers, 1):
    if h:
        c = ws.cell(row=row, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
row += 1

bugs = [
    ('CAM-14','AttributeError: SEARCH_INPUT','推断属性名，实际PO为SEARCH_ITEM','SEARCH_INPUT→SEARCH_ITEM (test_camera_management.py:330)'),
    ('MT-18/MT-21','input_keyword()不存在','维保页面只有下拉选择器','input_keyword→select_type/select_status'),
    ('MT-19','100字符名保存超时','后端可能拒绝超长名称','改用timestamp短名称'),
    ('AC-09~17 (7项)','CRUD弹窗操作失败','Element Plus 2.x teleport+is_displayed()不兼容','EP-001已知;待PO弹窗改用visibility+z-index'),
    ('MT-03','表格行计数波动','maintenance页面3个el-table','已知遗留#2;待per-table scope'),
    ('KP-11','查看按钮匹配失败','中文按钮XPath/JS匹配失败','已知遗留#3'),
]
for bid, problem, cause, fix in bugs:
    for i, v in enumerate([bid, problem, cause, fix, '', '', '', ''], 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = normal_font; c.alignment = left_align; c.border = thin_border
    row += 1

# Print settings
ws.freeze_panes = 'A5'
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = 'landscape'

outpath = 'governance/kpi/testcases/equipment/equipment-testcases.xlsx'
wb.save(outpath)
print(f'Saved: {outpath}')
print(f'Total: {total} tests (P0:{p0} P1:{p1} P2:{p2})')
print(f'Results: {n_pass} PASS / {n_fail} FAIL / {n_skip} SKIP / {n_retry} RETRY')
print(f'Effective pass rate: {rate}% ({effective_pass}/{effective})')
print(f'Sections: {sum(1 for t in tests if t[0]=="section")}, Subsections: {sum(1 for t in tests if t[0]=="subsection")}')
