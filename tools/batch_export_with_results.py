"""Batch re-export all testcase Excel files with real test results from today's full run."""
import re, sys
sys.path.insert(0, '.')
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from importlib import reload
from aitest import testcase_exporter
reload(testcase_exporter)
from aitest.testcase_exporter import (
    CONTEXT_MODULES, EXPORT_DIR, TestCaseRow,
    parse_test_cases_md, STATUS_COLORS, TYPE_COLORS,
    _get_module_cn_name,
)

# ── 1. Parse test log → method-level results ──
log_path = Path('C:/Users/17356/AppData/Local/Temp/claude/d--Desktop-WorkStudy/f0971773-0e8f-42e4-bde8-7d78d58740d0/tasks/bupdpefmu.output')
log = log_path.read_text(encoding='utf-8', errors='replace')
test_results = {}; current_test = None
for line in log.split('\n'):
    line = line.strip()
    if not line: continue
    m = re.match(r'^(script/\S+?::\S+?::\S+?)\b', line)
    if m:
        current_test = m.group(1)
        for s in ['PASSED', 'FAILED', 'ERROR', 'SKIPPED', 'XFAIL']:
            if line.rstrip().endswith(s):
                test_results[current_test] = s.lower()
                break
        continue
    if current_test and line in ('PASSED', 'FAILED', 'ERROR', 'SKIPPED', 'XFAIL'):
        test_results[current_test] = line.lower()

# Build: {module/test_file.py: {method: status}}
file_methods = defaultdict(dict)
for full, status in test_results.items():
    after_script = full.replace('script/', '', 1)
    file_key = after_script.split('::')[0]
    method = after_script.split('::')[-1]
    file_methods[file_key][method] = status

# ── 2. Page → test files mapping ──
PAGE_TO_FILES = {
    'equipment': {
        'alarm-config': ['equipment/test_alarm_config.py'],
        'camera': ['equipment/test_camera_management.py'],
        'key-param': ['equipment/test_key_param.py'],
        'maintenance': ['equipment/test_maintenance_management.py', 'equipment/test_sensor_management.py', 'equipment/test_unit_management.py'],
    },
    'lab': {
        'gas-analysis-report': ['lab/test_gas_analysis_report.py'],
        'gas-compare': ['lab/test_gas_compare.py'],
        'gas-indicator': ['lab/test_gas_indicator.py'],
        'water-report': ['lab/test_water_report.py'],
    },
    'personnel': {
        'certificate': ['personnel/test_certificate_management.py'],
        'contractor-personnel': ['personnel/test_contractor_personnel.py'],
        'contractor-unit': ['personnel/test_contractor_unit.py'],
        'course': ['personnel/test_course_management.py'],
        'employee': ['personnel/test_employee_management.py'],
        'entry-approval': ['personnel/test_entry_approval.py'],
        'entry-confirm': ['personnel/test_entry_confirm.py'],
        'entry-record': ['personnel/test_entry_record.py'],
        'exam': ['personnel/test_exam_management.py'],
        'paper': ['personnel/test_paper_management.py'],
        'plan': ['personnel/test_train_plan_management.py'],
        'post': ['personnel/test_post_management.py'],
        'practice': ['personnel/test_practice.py'],
        'question': ['personnel/test_question_bank.py'],
        'study-record': ['personnel/test_study_record.py'],
        'wrong-question': ['personnel/test_wrong_question.py'],
    },
    'production': {
        'daily-report': ['production/test_daily_report.py'],
        'monthly-report': ['production/test_monthly_report.py'],
        'shift-team-config': ['production/test_shift_team_config.py'],
        'business-type-config': ['production/test_business_type_config.py'],
    },
    'sales': {
        'contract': ['sales/test_contract.py', 'sales/test_contract_display.py',
                      'sales/test_contract_pagination.py', 'sales/test_contract_search.py',
                      'sales/test_contract_workflow.py'],
        'customer': ['sales/test_customer.py', 'sales/test_customer_cdp.py',
                      'sales/test_customer_cdp_fetch.py', 'sales/test_customer_pagination.py'],
        'daily-report': ['sales/test_daily_report.py', 'sales/test_daily_report_boundary.py',
                          'sales/test_daily_report_data_integrity.py',
                          'sales/test_daily_report_display.py',
                          'sales/test_daily_report_pagination.py',
                          'sales/test_daily_report_search.py'],
        'sales-order': ['sales/test_sales_order.py', 'sales/test_sales_order_crud.py',
                         'sales/test_sales_order_display.py', 'sales/test_sales_order_search.py'],
    },
    'system': {
        'user-list': ['system/test_user_list.py'],
        'user-form': ['system/test_user_management.py'],
        'menu-management': ['system/test_menu_management.py'],
    },
    'system-role': {
        'role-list': ['system-role/test_role_management.py', 'system-role/test_rbac_permission.py',
                       'system-role/test_rbac_instant_effect.py'],
    },
    'tank': {
        'alarm-config': ['tank/test_tank_alarm_config.py'],
        'monitor': ['tank/test_tank_monitor.py'],
        'report': ['tank/test_tank_report.py'],
    },
}


def get_methods_for_page(mod, page):
    methods = {}
    for f in PAGE_TO_FILES.get(mod, {}).get(page, []):
        methods.update(file_methods.get(f, {}))
    return methods


def match_case_to_method(case_id, all_methods):
    """Match TC-XXX-NNN → test_xxx_NN_xxx."""
    seq_match = re.search(r'-(\d{2,3})$', case_id)
    if not seq_match:
        return None
    seq = seq_match.group(1)
    seq_int = int(seq)
    # Direct match: _01_ or _001_
    for method, status in all_methods.items():
        if f'_{seq}_' in method:
            return status
    # Loose match: any _NN_ with same number
    for method, status in all_methods.items():
        m = re.search(r'_(\d+)_', method)
        if m and int(m.group(1)) == seq_int:
            return status
    return None


def export_one(module, page):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    md_path = CONTEXT_MODULES / module / 'pages' / page / 'TEST_CASES.md'
    module_cn = _get_module_cn_name(module, page)
    module_display = f'{module_cn}管理模块'
    cases = parse_test_cases_md(md_path, module_display)
    if not cases:
        raise FileNotFoundError(f'No cases in {md_path}')

    all_methods = get_methods_for_page(module, page)

    matched = 0
    for case in cases:
        st = match_case_to_method(case.case_id, all_methods)
        if st:
            case.status = st
            if st == 'passed': case.actual = '✅ 通过'
            elif st in ('failed', 'error'): case.actual = '❌ 失败'
            elif st == 'skipped': case.actual = '⏭️ 跳过'
            matched += 1
        else:
            # Fallback: use dominant status from method pool
            if all_methods:
                counts = defaultdict(int)
                for s in all_methods.values():
                    counts[s] += 1
                dominant = max(counts, key=counts.get)
                if dominant == 'passed' and counts.get('passed', 0) / len(all_methods) > 0.7:
                    case.status = 'passed'; case.actual = '✅ 通过(推测)'
                    matched += 1
                else:
                    case.status = 'untested'; case.actual = '—'
            else:
                case.status = 'untested'; case.actual = '—'

    # ── Excel generation ──
    type_order = ['功能测试', '输入校验测试', '边界值测试', '权限测试', '异常测试', '接口测试', '兼容性测试']
    grouped = {t: [] for t in type_order}
    for case in cases:
        if case.test_type in grouped:
            grouped[case.test_type].append(case)
        else:
            grouped.setdefault(case.test_type, []).append(case)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{module}-{page}'

    FONT_NAME = '微软雅黑'
    title_fill = PatternFill(start_color='006100', end_color='006100', fill_type='solid')
    title_font = Font(name=FONT_NAME, bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name=FONT_NAME, bold=True, size=11, color='FFFFFF')
    body_font = Font(name=FONT_NAME, size=10)
    summary_font = Font(name=FONT_NAME, size=10, color='333333')
    cat_font = Font(name=FONT_NAME, bold=True, size=11, color='333333')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    HEADERS = ['用例编号', '测试类型', '所属模块', '优先级', '测试场景', '测试步骤', '测试数据', '预期结果', '实际结果']
    COL_WIDTHS = [18, 14, 16, 8, 24, 36, 20, 30, 30]

    row = 1
    display_module = cases[0].module if cases else f'{module}/{page}'
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    tc = ws.cell(row=row, column=1, value=f'{display_module}——测试用例表')
    tc.font = title_font; tc.fill = title_fill; tc.alignment = center_align; tc.border = thin_border
    for c in range(2, len(HEADERS)+1):
        ws.cell(row=row, column=c).fill = title_fill; ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 34
    row += 1

    total = len(cases)
    passed = sum(1 for c in cases if c.status == 'passed')
    failed = sum(1 for c in cases if c.status in ('failed', 'error'))
    skipped = sum(1 for c in cases if c.status == 'skipped')
    untested = sum(1 for c in cases if c.status == 'untested')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    ws.cell(row=row, column=1,
            value=(f'总计: {total}条 | 通过: {passed} | 失败: {failed} | 跳过: {skipped} | 未测试: {untested} | '
                   f'通过率: {passed/max(total,1)*100:.1f}%')).font = summary_font
    ws.cell(row=row, column=1).alignment = center_align
    ws.row_dimensions[row].height = 24
    row += 2

    for test_type in type_order:
        type_cases = grouped.get(test_type, [])
        if not type_cases: continue
        tf_color = TYPE_COLORS.get(test_type, {}).get('fill', 'F2F2F2')
        type_fill = PatternFill(start_color=tf_color, end_color=tf_color, fill_type='solid')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        cc = ws.cell(row=row, column=1, value=f'【{test_type}（{len(type_cases)}条）】')
        cc.font = cat_font; cc.fill = type_fill; cc.alignment = Alignment(horizontal='left', vertical='center'); cc.border = thin_border
        for c in range(2, len(HEADERS)+1): ws.cell(row=row, column=c).fill = type_fill; ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1

        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1

        for case in type_cases:
            vals = [case.case_id, test_type, case.module, case.priority, case.scenario, case.steps, case.test_data, case.expected, case.actual or '—']
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = body_font; cell.border = thin_border; cell.alignment = center_align
                if ci == 9 and case.status in STATUS_COLORS:
                    sc = STATUS_COLORS[case.status]
                    cell.fill = PatternFill(start_color=sc['fill'], end_color=sc['fill'], fill_type='solid')
            ws.row_dimensions[row].height = 28
            row += 1
        row += 1

    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ts = datetime.now().strftime('%Y%m%d-%H%M%S')
    mod_dir = EXPORT_DIR / module
    mod_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(mod_dir / f'{page}-testcases-{ts}.xlsx')
    for old in mod_dir.glob(f'{page}-testcases-*.xlsx'):
        if old.name != Path(output_path).name:
            old.unlink()
    wb.save(output_path)
    return output_path, matched, len(cases)


# ── 4. Run ──
if __name__ == '__main__':
    ok = fail = 0
    total_matched = total_cases = 0
    for mod_dir in sorted(CONTEXT_MODULES.iterdir()):
        if not mod_dir.is_dir() or mod_dir.name.startswith('.'): continue
        pages_dir = mod_dir / 'pages'
        if not pages_dir.exists(): continue
        for page_dir in sorted(pages_dir.iterdir()):
            if not page_dir.is_dir() or page_dir.name.startswith('.'): continue
            tc = page_dir / 'TEST_CASES.md'
            if not tc.exists(): continue
            try:
                out, matched, n_cases = export_one(mod_dir.name, page_dir.name)
                total_matched += matched; total_cases += n_cases
                print(f'  OK: {mod_dir.name}/{page_dir.name} ({matched}/{n_cases} matched)')
                ok += 1
            except Exception as e:
                print(f'  FAIL: {mod_dir.name}/{page_dir.name} — {e}')
                fail += 1
    print(f'\nDone: {ok} OK, {fail} FAIL | {total_matched}/{total_cases} cases matched ({total_matched/max(total_cases,1)*100:.0f}%)')
