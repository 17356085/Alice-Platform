"""Generate 4 page-level Excel per spec: 1 page = 1 Excel, grouped by test category.
Section rows: merged, blue fill, bold, with count (e.g. "功能测试 (4条)").
Spec: governance/skills/execution/excel-exporter.md (Scenario C, 11-column)
Memory: excel-one-per-page + category-grouping
"""
import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO = Path(__file__).resolve().parent.parent
OUT_DIR = REPO / 'governance' / 'kpi' / 'reports' / 'equipment'
os.makedirs(OUT_DIR, exist_ok=True)

# ── Styles ──
HDR_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HDR_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')
INFO_FONT = Font(name='微软雅黑', size=9, color='666666')
BODY_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT = Font(name='微软雅黑', size=10, bold=True)
SECTION_FONT = Font(name='微软雅黑', size=11, bold=True)
SECTION_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
WRAP = Alignment(wrap_text=True, vertical='top', horizontal='left')
CENTER = Alignment(wrap_text=True, vertical='center', horizontal='center')
LEFT = Alignment(wrap_text=True, vertical='center', horizontal='left')
BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

P0_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
P1_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
P2_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
P_FILLS = {'P0': P0_FILL, 'P1': P1_FILL, 'P2': P2_FILL}

PASS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FAIL_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
SKIP_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
UNEXEC_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
STATUS_FILLS = {'PASS': PASS_FILL, 'FAIL': FAIL_FILL, 'SKIP': SKIP_FILL, '未执行': UNEXEC_FILL}

AUTO_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
MANUAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

HEADERS = ['用例编号', '用例标题', '优先级', '前置条件', '测试步骤',
           '测试数据', '预期结果', '实际状态', '耗时(s)', '错误信息', '自动化']
COL_WIDTHS = {'A': 18, 'B': 32, 'C': 8, 'D': 22, 'E': 40,
              'F': 18, 'G': 36, 'H': 10, 'I': 10, 'J': 38, 'K': 10}

PAGE_NAMES = {
    'alarm-config': ('设备报警配置', '设备管理'),
    'camera': ('摄像头管理', '设备管理'),
    'key-param': ('关键参数监控', '设备管理'),
    'maintenance': ('设备维保管理', '设备管理'),
}

# ── Test data: dict of page -> [(section_label, [cases]), ...]
# Each case: (cid, title, pri, pre, steps, data, expected, status, dur, err, auto)
PAGES = {}

# Helper: register case with all 11 fields, defaults for missing
def make_case(cid, title, pri, pre, steps, data, expected, status, dur='', err='', auto='✅'):
    return (cid, title, pri, pre, steps, data, expected, status, dur or '', err or '', auto)

# ════════════════════════════════════════════════════════════════
#  alarm-config
# ════════════════════════════════════════════════════════════════
ac = []
ac.append(('功能测试 (5条)', [
    make_case('AC-01','页面正常加载','P0','admin登录','导航至报警配置→等待渲染','—','统计卡片≥4张;统计非空','PASS',auto='✅'),
    make_case('AC-02','表格表头正确显示','P0','AC-01已通过','获取表头→校验9列','EXPECTED_TABLE_HEADER_SET','表头9列完整','PASS'),
    make_case('AC-03','搜索输入框可见','P0','页面已加载','检查搜索输入框','—','可见','PASS'),
    make_case('AC-04','新增按钮可见','P0','页面已加载','检查新增配置按钮','—','可见','PASS'),
    make_case('AC-05','统计卡片数据一致性','P0','页面已加载','获取4项统计','—','全部≥0','PASS'),
]))

ac.append(('搜索筛选测试 (3条)', [
    make_case('AC-06','按关键词搜索','P1','有数据','输入关键词→查询','测试报警下-1#','仅匹配记录','PASS'),
    make_case('AC-07','搜索无匹配结果','P1','已加载','输入不存在关键词→查询','zzzz_nonexistent','不崩溃','PASS'),
    make_case('AC-08','重置搜索条件','P1','已设条件','输入+搜索→重置','—','恢复全量','PASS'),
]))

ac.append(('分页测试 (2条) 🆕', [
    make_case('AC-18','分页组件可见','P1','已加载','检查.el-pagination','—','分页存在','PASS'),
    make_case('AC-19','pageSize默认值','P1','已加载','获取pageSize','—','pageSize=10','PASS'),
]))

ac.append(('权限测试 (2条) 🆕', [
    make_case('AC-20','新增按钮权限','P1','admin登录','检查新增按钮','—','可见','PASS'),
    make_case('AC-21','行操作按钮权限','P1','表格有数据','检查编辑/删除','—','≥1个','SKIP',err='表格为空'),
]))

ac.append(('边界值测试 (2条) 🆕', [
    make_case('AC-22','搜索特殊字符','P1','已加载','输入!@#$%→查询','!@#$%^&*()_+...','不崩溃','PASS'),
    make_case('AC-23','搜索超长关键词','P1','已加载','输入256字符→查询','a×256','不崩溃','PASS'),
]))

ac.append(('批量操作测试 (1条) 🆕', [
    make_case('AC-24','批量复选框','P2','已加载','检查checkbox','—','批量检测','PASS'),
]))

ac.append(('增删改弹窗测试 (7条) ⚠️已知Teleport问题', [
    make_case('AC-09','新增-必填字段','P0','admin+权限','新增→填必填→保存','autotest_req_XXXXXX','新增成功','FAIL',err='teleport+is_displayed(EP-001)'),
    make_case('AC-10','新增-全部字段','P1','AC-09条件','填全部→保存','name+type+level...','保存成功','FAIL',err='同上teleport'),
    make_case('AC-11','新增-取消','P1','AC-09条件','新增→填后取消','—','不入库','FAIL',err='同上'),
    make_case('AC-12','编辑规则','P1','表格有数据','编辑→修改→保存','—','更新成功','FAIL',err='同上'),
    make_case('AC-13','删除确认弹窗','P1','表格有数据','删除→确认弹窗','—','删除成功','FAIL',err='同上'),
    make_case('AC-16','阈值相等边界','P1','AC-09条件','上下限相等→保存','upper=lower','系统校验','FAIL',err='同上teleport'),
    make_case('AC-17','双击防重复提交','P2','AC-09条件','快速双击保存','—','仅提交一次','FAIL',err='click intercepted'),
]))

ac.append(('其他测试 (2条)', [
    make_case('AC-14','查看详情','P1','表格有数据','查看→详情弹窗','—','弹窗显示','SKIP',err='teleport'),
    make_case('AC-15','状态切换','P1','表格有数据','切换el-switch','—','切换成功','SKIP',err='el-switch不可交互'),
]))

ac.append(('接口测试 (2条)', [
    make_case('AC-API-01','设备列表API','P1','有token','GET /api/equipment/device/list','—','code=200,records>0','PASS',auto='✅'),
    make_case('AC-API-02','用户列表API','P1','有token','GET /api/system/user/list','—','code=200,records>0','PASS'),
]))

PAGES['alarm-config'] = ac

# ════════════════════════════════════════════════════════════════
#  camera
# ════════════════════════════════════════════════════════════════
cm = []
cm.append(('功能测试 (4条)', [
    make_case('CAM-01','页面正常加载','P0','admin登录','导航→检查卡片+网格','—','卡片≥4;统计非空','PASS'),
    make_case('CAM-02','统计卡片数据','P0','CAM-01通过','获取总数/在线/离线/故障','—','4项≥0','PASS'),
    make_case('CAM-03','监控卡片网格','P0','CAM-01通过','检查monitor-cell','—','≥0','PASS'),
    make_case('CAM-04','统计标签校验','P0','CAM-01通过','校验4标签','—','标签完整','PASS'),
]))

cm.append(('搜索筛选测试 (2条)', [
    make_case('CAM-05','关键词搜索','P1','已加载','输入关键词→搜索','罐区','≥0','PASS'),
    make_case('CAM-06','搜索无结果','P1','已加载','输入不存在关键词→搜索','zzzz_nonexistent','不崩溃','PASS'),
]))

cm.append(('分页测试 (2条)', [
    make_case('CAM-07','分页组件','P1','已加载','检查分页','—','total≥0','PASS'),
    make_case('CAM-08','翻页测试','P1','数据>1页','下一页→查卡片','—','翻页正常','SKIP',err='数据不足1页'),
]))

cm.append(('数据一致性测试 (2条)', [
    make_case('CAM-11','统计与分页一致','P1','已加载','对比stat vs page total','—','stat≥page','PASS'),
    make_case('CAM-12','搜索后分页重置','P1','数据>1页','翻页→搜索→查页码','—','重置到1','SKIP',err='不足1页'),
]))

cm.append(('卡片状态测试 (2条)', [
    make_case('CAM-09','卡片状态标签','P1','有卡片','获取状态','—','在线/离线/故障','SKIP',err='无卡片'),
    make_case('CAM-10','卡片操作按钮','P1','有卡片','检查操作区','—','≥1个','SKIP',err='无卡片'),
]))

cm.append(('弹窗测试 (1条)', [
    make_case('CAM-13','查看/预览弹窗','P2','有卡片','点击查看→弹窗','—','弹窗正常','SKIP',err='无卡片'),
]))

cm.append(('权限测试 (2条) 🆕', [
    make_case('CAM-14','搜索框权限','P1','admin登录','检查搜索框','—','可见','PASS',err='Bug:SEARCH_INPUT→SEARCH_ITEM'),
    make_case('CAM-15','卡片操作按钮权限','P1','有卡片','检查按钮区','—','存在','SKIP',err='无卡片'),
]))

cm.append(('边界值测试 (3条) 🆕', [
    make_case('CAM-16','搜索特殊字符','P1','已加载','输入!@#$%→搜索','!@#$%^&*()','不崩溃','PASS'),
    make_case('CAM-17','搜索超长关键词','P1','已加载','输入256字符→搜索','x×256','不崩溃','PASS'),
    make_case('CAM-18','空搜索关键词','P1','已加载','清空→搜索','(空)','正常返回','PASS'),
]))

cm.append(('可靠性测试 (1条) 🆕', [
    make_case('CAM-19','重复搜索稳定性','P2','已加载','3次不同关键词搜索','罐区,消防,(空)','3次正常','PASS'),
]))

PAGES['camera'] = cm

# ════════════════════════════════════════════════════════════════
#  key-param
# ════════════════════════════════════════════════════════════════
kp = []
kp.append(('功能测试 (3条)', [
    make_case('KP-01','页面正常加载','P0','admin登录','导航→查卡片+表格+分页','—','卡片≥2;表格;分页','PASS'),
    make_case('KP-02','统计卡片数据','P0','KP-01通过','获取4项统计','—','4项≥0','PASS'),
    make_case('KP-03','表头校验','P0','KP-01通过','获取表头','—','≥5列','PASS'),
]))

kp.append(('搜索筛选测试 (3条)', [
    make_case('KP-04','关键词自动过滤','P0','已加载','输入关键词→刷新','温度/压力/传感器','行数变化','PASS'),
    make_case('KP-05','搜索无匹配','P1','已加载','输入不存在关键词','zzzz_nonexistent','行数变化','PASS'),
    make_case('KP-06','重置恢复全量','P1','已设搜索','输入→重置','—','total=全量','PASS'),
]))

kp.append(('数据校验测试 (1条)', [
    make_case('KP-10','状态逻辑一致性','P0','表格有数据','前5条→校验状态vs阈值','—','状态与阈值一致','SKIP',err='数据校验未完成'),
]))

kp.append(('弹窗测试 (1条)', [
    make_case('KP-11','查看参数详情','P1','表格有数据','查看→弹窗','—','弹窗显示','SKIP',err='Unicode匹配',auto='⏸️'),
]))

kp.append(('分页测试 (2条) 🆕', [
    make_case('KP-12','分页组件+总数','P1','已加载','检查分页','—','total≥0','PASS'),
    make_case('KP-13','pageSize默认','P1','已加载','获取pageSize','—','=10','PASS'),
]))

kp.append(('权限测试 (1条) 🆕', [
    make_case('KP-14','行操作按钮权限','P1','表格有数据','检查第一行按钮','—','≥1个','PASS'),
]))

kp.append(('边界值测试 (2条) 🆕', [
    make_case('KP-15','搜索特殊字符','P1','已加载','输入!@#$%→刷新','!@#$%^&*()','不崩溃','PASS'),
    make_case('KP-16','搜索超长关键词','P1','已加载','输入256字符→刷新','x×256','不崩溃','PASS'),
]))

kp.append(('可靠性测试 (1条) 🆕', [
    make_case('KP-17','重复搜索稳定性','P2','已加载','3次:重置→输入→刷新','温度×3','3次正常','PASS'),
]))

PAGES['key-param'] = kp

# ════════════════════════════════════════════════════════════════
#  maintenance
# ════════════════════════════════════════════════════════════════
mt = []
mt.append(('功能测试 (4条)', [
    make_case('MT-01','页面正常加载','P0','admin登录','导航→检查URL+标题+表头','—','URL含maintenance','PASS'),
    make_case('MT-02','表头正确','P0','MT-01通过','检查表头','—','≥9列','PASS'),
    make_case('MT-03','表格有数据','P0','有维保计划','获取行数','—','>0','FAIL',err='3表跨表行计数',auto='🔄'),
    make_case('MT-04','分页组件','P0','已加载','检查分页','—','total>0','PASS'),
]))

mt.append(('搜索筛选测试 (3条)', [
    make_case('MT-05','按类型搜索','P1','已加载','选择日检→搜索','日检','正常','PASS'),
    make_case('MT-06','按状态搜索','P1','已加载','选择待执行→搜索','待执行','正常','PASS'),
    make_case('MT-07','重置恢复全量','P1','已设搜索','选类型+搜索→重置','—','total≥原','PASS'),
]))

mt.append(('分页测试 (2条)', [
    make_case('MT-08','分页下一页','P1','数据>1页','下一页→查行数','—','有数据','PASS'),
    make_case('MT-09','切换每页条数','P1','数据>10','切换20→检查→恢复10','—','行数≤20','PASS'),
]))

mt.append(('新增编辑测试 (5条)', [
    make_case('MT-10','新增计划-保存','P0','admin+权限','新增→填写→保存','AUTO_测试计划','Toast成功','PASS'),
    make_case('MT-11','新增-取消','P0','MT-10条件','新增→填写→取消','—','不入库','PASS'),
    make_case('MT-12','新增-必填校验','P1','MT-10条件','新增→直接保存','—','校验提示','PASS'),
    make_case('MT-13','编辑-改名称','P0','有数据','编辑→改名→保存→还原','原→原_改→原','更新+还原','PASS'),
    make_case('MT-14','编辑-取消','P1','有数据','编辑→改名→取消','—','名称不变','PASS'),
]))

mt.append(('功能测试 (1条)', [
    make_case('MT-15','生成维保任务','P1','有数据','生成任务→Toast','—','有反馈','PASS'),
]))

mt.append(('权限测试 (2条)', [
    make_case('MT-16','新增按钮(admin)','P1','admin登录','检查新增','—','可见','PASS'),
    make_case('MT-17','编辑按钮可见','P1','有编码','检查编辑','—','可见','PASS'),
]))

mt.append(('边界值测试 (2条) 🆕', [
    make_case('MT-18','下拉类型组合','P1','已加载','遍历日检/周检/月检','日检,周检,月检','每种正常','PASS',err='Bug:input_keyword→select_type'),
    make_case('MT-19','计划名称标识','P1','MT-10条件','新增→timestamp名','AUTOTEST_{ts}','成功','PASS',err='Bug:100char→timestamp'),
]))

mt.append(('批量操作测试 (1条) 🆕', [
    make_case('MT-20','批量复选框','P1','已加载','检查checkbox','—','检测','PASS'),
]))

mt.append(('可靠性测试 (1条) 🆕', [
    make_case('MT-21','重复搜索稳定性','P2','已加载','3次:重置→选类型→搜索','日检,待执行','3次正常','PASS',err='Bug:input_keyword→select'),
]))

PAGES['maintenance'] = mt

# ════════════════════════════════════════════════════════════════
#  Generate one Excel per page
# ════════════════════════════════════════════════════════════════
total_all = 0; pass_all = 0

for page_slug, sections in PAGES.items():
    page_name, module_name = PAGE_NAMES[page_slug]

    # Flatten to get total stats
    all_cases = []
    for sec_label, cases in sections:
        all_cases.extend(cases)
    total = len(all_cases)
    n_pass = sum(1 for t in all_cases if t[7] == 'PASS')
    n_fail = sum(1 for t in all_cases if t[7] == 'FAIL')
    n_skip = sum(1 for t in all_cases if t[7] == 'SKIP')
    rate = f'{n_pass/max(total,1)*100:.1f}%'
    p0 = sum(1 for t in all_cases if t[2]=='P0')
    p1 = sum(1 for t in all_cases if t[2]=='P1')
    p2 = sum(1 for t in all_cases if t[2]=='P2')
    total_all += total; pass_all += n_pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = page_slug

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.cell(row=1, column=1, value=f'{module_name} — {page_name} 测试报告').font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # Row 2: Info
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    info = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")} | 页面: {page_name}({page_slug}) | 总: {total} (P0:{p0} P1:{p1} P2:{p2}) | 通过: {n_pass} | 失败: {n_fail} | 跳过: {n_skip} | 通过率: {rate}'
    ws.cell(row=2, column=1, value=info).font = INFO_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    # Row 4: Headers
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[4].height = 22

    # Row 5+: Sections + data rows
    row = 5
    for sec_label, cases in sections:
        sec_count = len(cases)
        label = f'{sec_label}' if '(' in sec_label else f'{sec_label} ({sec_count}条)'

        # Section header row (merged, bold, light blue fill)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        cell = ws.cell(row=row, column=1, value=f'  {label}')
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = LEFT
        for cc in range(1, 12):
            ws.cell(row=row, column=cc).border = BORDER
            ws.cell(row=row, column=cc).fill = SECTION_FILL
        ws.row_dimensions[row].height = 24
        row += 1

        for t in cases:
            cid, title, pri, pre, steps, data, expected, status, dur, err, auto = t
            values = [cid, title, pri, pre, steps, data, expected, status, dur, err, auto]
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.font = BODY_FONT
                cell.alignment = CENTER
                cell.border = BORDER

            # Priority fill (col C)
            if pri in P_FILLS:
                ws.cell(row=row, column=3).fill = P_FILLS[pri]

            # Status fill (col H)
            if status in STATUS_FILLS:
                sc = ws.cell(row=row, column=8)
                sc.fill = STATUS_FILLS[status]
                if status == 'FAIL':
                    sc.font = BOLD_FONT

            # Auto fill (col K)
            if auto in ('✅', '✓', '已自动化'):
                ws.cell(row=row, column=11).fill = AUTO_FILL
            elif auto in ('❌', '不适合'):
                ws.cell(row=row, column=11).fill = MANUAL_FILL

            ws.row_dimensions[row].height = 28
            row += 1

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = 'A5'

    fname = f'测试报告-equipment-{page_slug}.xlsx'
    outpath = OUT_DIR / fname
    wb.save(outpath)
    print(f'[DONE] {fname:50} {total:2} cases | PASS={n_pass:2} FAIL={n_fail:2} SKIP={n_skip:2} | {rate} | {len(sections)} sections')

# ── Generate .md test case files (SOP Phase 3 deliverable) ──
MD_OUT = REPO / 'governance' / 'kpi' / 'testcases' / 'equipment'
os.makedirs(MD_OUT, exist_ok=True)
today_str = datetime.now().strftime('%Y-%m-%d')

for page_slug, sections in PAGES.items():
    page_name, module_name = PAGE_NAMES[page_slug]
    lines = []
    lines.append(f'# {module_name} — {page_name} 测试用例')
    lines.append('')
    lines.append(f'> 模块: equipment | 页面: {page_slug} | 生成: {today_str}')
    lines.append('')

    # Flatten stats
    all_cases_md = []
    for _, cases in sections:
        all_cases_md.extend(cases)
    total_md = len(all_cases_md)
    n_pass_md = sum(1 for t in all_cases_md if t[7] == 'PASS')
    n_fail_md = sum(1 for t in all_cases_md if t[7] == 'FAIL')
    n_skip_md = sum(1 for t in all_cases_md if t[7] == 'SKIP')

    lines.append(f'| 指标 | 数值 |')
    lines.append(f'|------|:---:|')
    lines.append(f'| 用例总数 | {total_md} |')
    lines.append(f'| 通过 | {n_pass_md} |')
    lines.append(f'| 失败 | {n_fail_md} |')
    lines.append(f'| 跳过 | {n_skip_md} |')
    lines.append(f'| 通过率 | {n_pass_md/max(total_md,1)*100:.1f}% |')
    lines.append('')

    for sec_label, cases in sections:
        count = len(cases)
        label = f'{sec_label}'
        lines.append(f'## {label}')
        lines.append('')
        lines.append(f'| 用例编号 | 用例标题 | 优先级 | 前置条件 | 测试步骤 | 测试数据 | 预期结果 | 实际状态 |')
        lines.append(f'|---------|---------|--------|---------|---------|---------|---------|---------|')
        for t in cases:
            cid, title, pri, pre, steps, data, expected, status, dur, err, auto = t
            pre_short = (pre or '—')[:40].replace('\n',' ')
            steps_short = (steps or '—')[:50].replace('\n','→')
            expected_short = (expected or '—')[:40]
            data_short = (data or '—')[:25]
            status_disp = {'PASS':'✅','FAIL':'❌','SKIP':'⏭'}.get(status, status)
            lines.append(f'| {cid} | {title} | **{pri}** | {pre_short} | {steps_short} | {data_short} | {expected_short} | {status_disp} |')
        lines.append('')

    fname = f'testcases-equipment-{page_slug}.md'
    with open(MD_OUT / fname, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f'[MD] {fname:50} {total_md} cases | {len(sections)} sections')

print(f'---')
print(f'[ALL] {total_all} cases | PASS={pass_all} | overall={pass_all/max(total_all,1)*100:.1f}%')
print(f'[Excel] 4 files → {OUT_DIR}')
print(f'[.md]   4 files → {MD_OUT}')
